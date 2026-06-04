from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import redirect, render, get_object_or_404
from django.utils import timezone

from apps.core.models import Company, Worker, Brigade
from apps.public.models import Banner, PublicPost, HeroBanner, SliderAd, SubBanner
from apps.accounts.models import AdminGroup, UserCompanyProfile


def home(request):
    if request.user.is_authenticated:
        return redirect("personal_profile")
    return redirect("/public/")


def get_display_name(user):
    if not user.is_authenticated:
        return ""
    full_name = user.get_full_name()
    if full_name:
        return full_name
    return user.username


def _build_public_context(request, tab_override=None):
    user = request.user
    tab = (tab_override or request.GET.get("tab") or "home").strip()

    is_admin_like = (
        user.is_authenticated
        and (
            user.is_superuser
            or user.is_staff
            or user.groups.filter(name__in=["ADMIN_FULL"]).exists()
        )
    )

    banners_qs = Banner.objects.filter(is_active=True).order_by("sort_order", "-created_at", "-id")
    banners = list(banners_qs[:50])

    posts_qs = PublicPost.objects.filter(is_published=True).select_related("author").order_by("-created_at", "-id")

    featured_admin_post = (
        posts_qs.filter(author__is_superuser=True).first()
        or posts_qs.filter(author__is_staff=True).first()
        or posts_qs.filter(author__groups__name__in=["ADMIN_FULL"]).distinct().first()
    )

    chat_stream = list(posts_qs[:200])
    hero_banner = HeroBanner.objects.filter(is_active=True).order_by("sort_order", "-created_at").first()
    slider_ads = list(SliderAd.objects.filter(is_active=True).order_by("sort_order", "-created_at")[:10])
    sub_banner = SubBanner.objects.filter(is_active=True).order_by("sort_order", "-created_at").first()

    contact = {
        "title": "Холбоо барих",
        "subtitle": "Санал хүсэлт, хамтын ажиллагаа, зар сурталчилгаа зэрэгт бидэнтэй холбогдоорой.",
        "phone": "+976 9911-2233",
        "email": "info@construction.mn",
        "address": "Улаанбаатар хот, СБД, 1-р хороо",
        "hours": "Даваа–Баасан 09:00–18:00",
        "facebook": "https://facebook.com/",
        "website": "https://example.com",
    }

    companies_count = Company.objects.count()
    workers_count = Worker.objects.count()
    brigades_count = Brigade.objects.count()
    recent_companies = list(Company.objects.order_by('-id')[:6])
    return {
        "tab": tab,
        "contact": contact,
        "banners": banners,
        "featured_admin_post": featured_admin_post,
        "chat_stream": chat_stream,
        "is_admin_like": is_admin_like,
        "can_post": user.is_authenticated,
        "display_name": get_display_name(user),
        "hero_banner": hero_banner,
        "slider_ads": slider_ads,
        "sub_banner": sub_banner,
    }


def public_home(request):
    user = request.user
    post_error = ""
    if request.method == "POST" and request.POST.get("action") == "new_post":
        if not user.is_authenticated:
            return redirect("/login/")
        title = (request.POST.get("post_title") or "").strip()
        body = (request.POST.get("post_body") or "").strip()
        if not body:
            post_error = "Чатын бичвэр хоосон байна."
        else:
            if not title:
                title = timezone.now().strftime("Чат %Y-%m-%d %H:%M")
            PublicPost.objects.create(author=user, title=title, body=body, is_published=True)
            return redirect("/public/")

    context = _build_public_context(request)
    context["post_error"] = post_error
    return render(request, "registry/public_home.html", context)


def public_login(request):
    if request.user.is_authenticated:
        next_url = request.GET.get("next", "/public/")
        return redirect(next_url)

    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            auth_login(request, form.get_user())
            next_url = request.POST.get("next", request.GET.get("next", "/public/"))
            return redirect(next_url)

    next_url = request.GET.get("next", "")
    return render(request, "registration/login.html", {"form": form, "next": next_url})

# =====================================================
# ✅ ШИНЭ: Компанийн нийтийн хуудас
# =====================================================
def company_profile(request, slug):
    company = get_object_or_404(Company, slug=slug)

    workers_qs = Worker.objects.filter(company=company).order_by("responsible_role", "first_name")
    brigades_qs = Brigade.objects.filter(companies=company).distinct()

    is_auth = request.user.is_authenticated

    is_owner = False
    if is_auth:
        prof = getattr(request.user, "company_profile", None)
        if prof and prof.company_id == company.id:
            is_owner = True
        if request.user.is_superuser or request.user.is_staff:
            is_owner = True

    context = {
        "company": company,
        "workers": list(workers_qs) if is_auth else list(workers_qs[:3]),
        "workers_total": workers_qs.count(),
        "brigades": list(brigades_qs) if is_auth else list(brigades_qs[:2]),
        "brigades_total": brigades_qs.count(),
        "is_auth": is_auth,
        "is_owner": is_owner,
        "display_name": get_display_name(request.user),
    }
    return render(request, "registry/company_profile.html", context)


@login_required
def personal_profile(request):
    user = request.user

    if request.session.get("profile_unlocked") is True:
        profile = UserCompanyProfile.objects.filter(user=user).select_related("company").first()
        company = profile.company if profile else None
        from apps.public.models import Ad
        my_ads = list(Ad.objects.filter(author=user).order_by("-created_at")[:10])
        return render(request, "registry/profile.html", {"company": company, "my_ads": my_ads})

    error = ""
    if request.method == "POST":
        password = (request.POST.get("password") or "").strip()
        if not password:
            error = "Нууц үгээ оруулна уу."
        else:
            authed = authenticate(request, username=user.get_username(), password=password)
            if authed is None:
                error = "Нууц үг буруу байна."
            else:
                request.session["profile_unlocked"] = True
                return redirect("personal_profile")

    return render(request, "registry/profile.html", {"lock_mode": True, "error": error})


def auth_facebook(request):
    return render(request, "registry/auth_stub.html", {"provider": "Facebook"})


def auth_emongolia(request):
    return render(request, "registry/auth_stub.html", {"provider": "e-Mongolia"})


def auth_bank(request):
    return render(request, "registry/auth_stub.html", {"provider": "Банкны код"})
def register_view(request):
    from django.contrib.auth.models import User
    from django.contrib.auth import login as auth_login

    errors = ""
    if request.method == "POST":
        reg_type = request.POST.get("reg_type", "company")
        username = (request.POST.get("username") or "").strip()
        email = (request.POST.get("email") or "").strip()
        password1 = request.POST.get("password1", "")
        password2 = request.POST.get("password2", "")

        if not username:
            errors = "Хэрэглэгчийн нэр оруулна уу."
        elif User.objects.filter(username=username).exists():
            errors = "Энэ хэрэглэгчийн нэр аль хэдийн бүртгэлтэй байна."
        elif password1 != password2:
            errors = "Нууц үг таарахгүй байна."
        elif len(password1) < 6:
            errors = "Нууц үг хамгийн багадаа 6 тэмдэгт байх ёстой."
        else:
            user = User.objects.create_user(username=username, email=email, password=password1)
            if reg_type == "company":
                company_name = (request.POST.get("company_name") or "").strip()
                if company_name:
                    from apps.core.models import Company
                    company, _ = Company.objects.get_or_create(name=company_name)
                    from apps.accounts.models import UserCompanyProfile
                    UserCompanyProfile.objects.create(user=user, company=company)
            auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect("/public/")

    return render(request, "registration/register.html", {"errors": errors})

def ad_create(request):
    from apps.public.models import Ad
    from django.utils import timezone
    import datetime

    if not request.user.is_authenticated:
        return redirect("/login/?next=/ads/create/")

    errors = {}
    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        category = (request.POST.get("category") or "other").strip()
        description = (request.POST.get("description") or "").strip()
        price_str = (request.POST.get("price") or "").strip()
        price_type = (request.POST.get("price_type") or "negotiable").strip()
        city = (request.POST.get("city") or "UB").strip()
        district = (request.POST.get("district") or "").strip()
        contact_name = (request.POST.get("contact_name") or "").strip()
        contact_phone = (request.POST.get("contact_phone") or "").strip()
        contact_email = (request.POST.get("contact_email") or "").strip()

        if not title:
            errors["title"] = "Гарчиг заавал оруулна уу."
        if not contact_phone:
            errors["contact_phone"] = "Утасны дугаар заавал оруулна уу."

        if not errors:
            price = None
            if price_str:
                try:
                    price = float(price_str.replace(",", "").replace(" ", ""))
                except ValueError:
                    pass

            material_subcategory = (request.POST.get("material_subcategory") or "").strip()
            material_item = (request.POST.get("material_item") or "").strip()
            price_unit = (request.POST.get("price_unit") or "negotiable").strip()
            ad = Ad.objects.create(
                author=request.user,
                category=category,
                material_subcategory=material_subcategory,
                material_item=material_item,
                price_unit=price_unit,
                title=title,
                description=description,
                price=price,
                price_type=price_type,
                city=city,
                district=district,
                contact_name=contact_name,
                contact_phone=contact_phone,
                contact_email=contact_email,
                status="active",
                expires_at=timezone.now() + datetime.timedelta(days=30),
            )
            if request.FILES.get("image1"):
                ad.image1 = request.FILES["image1"]
            if request.FILES.get("image2"):
                ad.image2 = request.FILES["image2"]
            if request.FILES.get("image3"):
                ad.image3 = request.FILES["image3"]
            ad.save()
            return redirect("/ads/")

    return render(request, "registry/ad_create.html", {
        "errors": errors,
        "post_data": request.POST,
        "display_name": get_display_name(request.user),
    })


def ad_list(request):
    import json, os
    from apps.public.models import Ad
    category = request.GET.get("cat", "")
    subcat = request.GET.get("subcat", "")
    item_raw = request.GET.get("item", "")
    q = request.GET.get("q", "")

    # subcat__item форматаар задлах
    item = ""
    if item_raw and "__" in item_raw:
        parts = item_raw.split("__", 1)
        if not subcat:
            subcat = parts[0]
        item = parts[1]
    elif item_raw:
        item = item_raw

    ads = Ad.objects.filter(status="active").order_by("-created_at")
    if category:
        ads = ads.filter(category=category)
    if subcat:
        if category == "material":
            ads = ads.filter(material_subcategory=subcat)
        elif category == "house":
            ads = ads.filter(house_location_type=subcat)
    if item:
        if category == "material":
            ads = ads.filter(material_item=item)
        elif category == "house":
            if subcat == "rooms":
                ads = ads.filter(house_rooms=item)
            elif subcat in ("ub", "province"):
                ads = ads.filter(house_location=item)
            elif subcat == "type":
                ads = ads.filter(house_type=item)
    if q:
        ads = ads.filter(title__icontains=q)

    SUBCAT_LABELS = {
        "foundation": "Барилгын үндсэн хийц",
        "interior": "Засал чимэглэл",
        "outdoor": "Гадна тохижилт",
        "plumbing": "Сан, халаалт",
        "electrical": "Цахилгаан, холбоо",
        "machinery": "Машин, тоног",
        "furniture": "Тавилга",
        "software": "Программ, ном",
        "safety": "ХАБЭА",
        "rooms": "Өрөөний тоо",
        "ub": "Улаанбаатар",
        "province": "Орон нутаг",
        "type": "Зарын төрөл",
    }

    item_choices = []
    item_label = ""
    try:
        json_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "material_items.json")
        all_items = json.load(open(json_path, encoding="utf-8"))
        if subcat and subcat in all_items:
            item_choices = list(all_items[subcat].items())
            if item:
                item_label = all_items[subcat].get(item, "")
    except:
        pass

    return render(request, "registry/ad_list.html", {
        "ads": ads[:100],
        "category": category,
        "subcat": subcat,
        "item": item,
        "subcat_label": SUBCAT_LABELS.get(subcat, ""),
        "item_label": item_label,
        "item_choices": item_choices,
        "q": q,
        "display_name": get_display_name(request.user),
    })


def ad_detail(request, pk):
    from apps.public.models import Ad
    ad = Ad.objects.filter(pk=pk, status="active").first()
    if not ad:
        from django.http import Http404
        raise Http404
    ad.views += 1
    ad.save(update_fields=["views"])
    related = Ad.objects.filter(category=ad.category, status="active").exclude(pk=pk)[:4]
    return render(request, "registry/ad_detail.html", {
        "ad": ad,
        "related": related,
        "display_name": get_display_name(request.user),
    })


def news_list(request):
    from apps.public.models import PublicPost
    q = request.GET.get("q", "")
    posts = PublicPost.objects.filter(is_published=True).select_related("author").order_by("-created_at")
    if q:
        posts = posts.filter(title__icontains=q)
    return render(request, "registry/news_list.html", {
        "posts": posts[:50],
        "q": q,
        "display_name": get_display_name(request.user),
    })

def news_detail(request, pk):
    from apps.public.models import PublicPost
    post = PublicPost.objects.filter(pk=pk, is_published=True).select_related("author").first()
    if not post:
        from django.http import Http404
        raise Http404
    related = PublicPost.objects.filter(is_published=True).exclude(pk=pk).order_by("-created_at")[:4]
    return render(request, "registry/news_detail.html", {
        "post": post,
        "related": related,
        "display_name": get_display_name(request.user),
    })


def ad_delete(request, pk):
    from apps.public.models import Ad
    if not request.user.is_authenticated:
        return redirect("/login/")
    ad = Ad.objects.filter(pk=pk, author=request.user).first()
    if not ad:
        from django.http import Http404
        raise Http404
    if request.method == "POST":
        ad.delete()
        return redirect("/profile/")
    return render(request, "registry/ad_delete_confirm.html", {"ad": ad})


def tender_list(request):
    from apps.public.models import Tender
    from django.db.models import Count
    q = request.GET.get("q", "")
    cat = request.GET.get("cat", "")
    tenders = Tender.objects.order_by("-created_at")
    if q:
        tenders = tenders.filter(title__icontains=q) | tenders.filter(organization__icontains=q)
    if cat:
        tenders = tenders.filter(category=cat)
    counts = {r["category"]: r["n"] for r in
              Tender.objects.values("category").annotate(n=Count("id"))}
    return render(request, "registry/tender_list.html", {
        "tenders": tenders[:100],
        "q": q,
        "cat": cat,
        "total": Tender.objects.count(),
        "counts": counts,
        "display_name": get_display_name(request.user),
    })



def get_material_price(category, name_contains):
    from apps.public.models import MaterialPrice
    items = MaterialPrice.objects.filter(
        is_active=True,
        category=category,
        name__icontains=name_contains
    )
    if items.exists():
        p = items.first()
        return int((p.price_min + p.price_max) / 2)
    return 0

def calculate_budget_norm(data):
    floors = int(str(data.get("floors", 1)).replace("+",""))
    length = float(data.get("length", 10))
    width = float(data.get("width", 10))
    ceiling_h = float(str(data.get("ceiling_height", 2.7)).replace("+",""))

    floor_area = length * width
    total_area = floor_area * floors
    perimeter = 2 * (length + width)
    wall_height = ceiling_h + 0.3
    outer_wall_area = perimeter * wall_height * floors
    inner_wall_area = total_area * 0.35
    roof_area = floor_area * 1.15
    net_wall_area = round(outer_wall_area * 0.85, 1)

    wall_mat = data.get("wall_material", "Мак блок")
    quality = data.get("quality", "дунд").lower()
    foundation_type = data.get("foundation_type", "Шугаман суурь")
    foundation_depth = float(str(data.get("foundation_depth", 2.5)).replace("+","").replace("м",""))
    roof_type = data.get("roof_type", "")
    insulation = data.get("insulation", "")
    facade = data.get("facade", "Шавар штукатур")

    quality_coef = {"эконом": 0.8, "дунд": 1.0, "премиум": 1.35}.get(quality, 1.0)

    materials, labor, transport, other = [], [], [], []

    # 1. СУУРЬ
    if "хавтан" in foundation_type.lower():
        fv = round(floor_area * 0.3, 1)
        rebar_kg = floor_area * 25
        cg = "М300"
    elif "гадсан" in foundation_type.lower() or "нил" in foundation_type.lower():
        fv = round(floor_area * 0.15, 1)
        rebar_kg = floor_area * 20
        cg = "М300"
    else:
        fv = round(perimeter * foundation_depth * 0.5, 1)
        rebar_kg = fv * 80
        cg = "М250"

    rt = round(rebar_kg / 1000, 2)
    cp = get_material_price("mat_cement", "М250") or 270000
    rp = get_material_price("mat_rebar", "A III (d12") or 2500000
    sp = get_material_price("mat_sand", "Дайрга") or 37000
    ep = get_material_price("labor_general", "Газар шорооны") or 27000
    cwp = get_material_price("labor_general", "Бетон цутгалт") or 170000
    rwp = get_material_price("labor_general", "Арматур угсралт") or 1200000

    materials += [
        {"name": f"Бетон зуурмаг {cg} — суурь", "unit": "м³", "qty": fv, "unit_price": cp, "total": round(fv*cp)},
        {"name": "Арматур A III — суурь", "unit": "тонн", "qty": rt, "unit_price": rp, "total": round(rt*rp)},
        {"name": "Дайрга — суурийн доор", "unit": "м³", "qty": round(floor_area*0.1,1), "unit_price": sp, "total": round(floor_area*0.1*sp)},
    ]
    labor += [
        {"name": "Газар ухалт", "unit": "м³", "qty": round(fv*1.3,1), "unit_price": ep, "total": round(fv*1.3*ep)},
        {"name": "Бетон цутгалт — суурь", "unit": "м³", "qty": fv, "unit_price": cwp, "total": round(fv*cwp)},
        {"name": "Арматур угсралт — суурь", "unit": "тонн", "qty": rt, "unit_price": rwp, "total": round(rt*rwp)},
    ]

    # 2. ГАДНА ХАНА
    if "мак блок" in wall_mat.lower():
        bq = round(net_wall_area * 16)
        bp = get_material_price("mat_brick", "Мак блок (25") or 8500
        gp = get_material_price("mat_interior", "Блокны цавуу") or 15000
        wp = get_material_price("labor_general", "Блокон хана өрөх /25") or 25000
        materials += [
            {"name": "Мак блок (25см) — гадна хана", "unit": "ш", "qty": bq, "unit_price": bp, "total": bq*bp},
            {"name": "Блокны цавуу", "unit": "кг", "qty": round(net_wall_area*2), "unit_price": gp, "total": round(net_wall_area*2*gp)},
        ]
        labor.append({"name": "Мак блок хана өрөх", "unit": "м²", "qty": net_wall_area, "unit_price": wp, "total": round(net_wall_area*wp)})
    elif "тоосго" in wall_mat.lower():
        bq = round(net_wall_area * 51)
        bp = get_material_price("mat_brick", "Улаан тоосго") or 515
        wp = get_material_price("labor_general", "Тоосгон хана өрөх /25") or 25000
        materials.append({"name": "Улаан тоосго — гадна хана", "unit": "ш", "qty": bq, "unit_price": bp, "total": bq*bp})
        labor.append({"name": "Тоосгон хана өрөх", "unit": "м²", "qty": net_wall_area, "unit_price": wp, "total": round(net_wall_area*wp)})
    elif "бетон" in wall_mat.lower():
        wv = round(net_wall_area * 0.2, 1)
        cp2 = get_material_price("mat_cement", "М300") or 285000
        wp = get_material_price("labor_general", "Гадна бетон хананы") or 130000
        materials.append({"name": "Бетон зуурмаг М300 — хана", "unit": "м³", "qty": wv, "unit_price": cp2, "total": round(wv*cp2)})
        labor.append({"name": "Бетон хана цутгалт", "unit": "м³", "qty": wv, "unit_price": wp, "total": round(wv*wp)})

    # 3. ДУЛААЛГА
    if insulation and "байхгүй" not in insulation.lower():
        if "шилэн хөвөн" in insulation.lower():
            ip = get_material_price("mat_insulation", "Шилэн хөвөн (100") or 18500
        elif "хөөсөнцөр" in insulation.lower():
            ip = get_material_price("mat_insulation", "Хөөсөнцөр") or 8000
        elif "базальт" in insulation.lower():
            ip = get_material_price("mat_insulation", "Базальт") or 15000
        elif "xps" in insulation.lower():
            ip = get_material_price("mat_insulation", "XPS") or 21000
        else:
            ip = 15000
        iwp = get_material_price("labor_special", "Дулаалга хийх") or 11500
        materials.append({"name": f"Дулаалга ({insulation})", "unit": "м²", "qty": net_wall_area, "unit_price": ip, "total": round(net_wall_area*ip)})
        labor.append({"name": "Дулаалга хийх", "unit": "м²", "qty": net_wall_area, "unit_price": iwp, "total": round(net_wall_area*iwp)})

    # 4. ГАДНА ФАСАД
    plaster_work = get_material_price("labor_special", "Шавардлага") or 23500
    if "тоосго" in facade.lower() or "клинкер" in facade.lower():
        fap = get_material_price("mat_brick", "Өнгөлгөөний тоосго") or 3000
        faq = round(net_wall_area * 51)
        materials.append({"name": "Өнгөлгөөний тоосго — фасад", "unit": "ш", "qty": faq, "unit_price": fap, "total": faq*fap})
        labor.append({"name": "Өнгөлгөөний тоосго өрөх", "unit": "м²", "qty": net_wall_area, "unit_price": 35000, "total": round(net_wall_area*35000)})
    else:
        paint_p = 100000
        paint_q = round(net_wall_area / 8)
        materials += [
            {"name": "Цемент — фасадын шавардлага", "unit": "уут", "qty": round(net_wall_area*0.12), "unit_price": 41500, "total": round(net_wall_area*0.12*41500)},
            {"name": "Гадна фасадын будаг", "unit": "сав", "qty": paint_q, "unit_price": paint_p, "total": paint_q*paint_p},
        ]
        labor.append({"name": "Гадна ханын шавардлага", "unit": "м²", "qty": net_wall_area, "unit_price": plaster_work, "total": round(net_wall_area*plaster_work)})

    # 5. ХУЧИЛТ
    sa = floor_area * (floors-1) if floors > 1 else floor_area
    sv = round(sa * 0.2, 1)
    srt = round(sa * 12 / 1000, 2)
    materials += [
        {"name": "Бетон зуурмаг М250 — хучилт", "unit": "м³", "qty": sv, "unit_price": cp, "total": round(sv*cp)},
        {"name": "Арматур — хучилт", "unit": "тонн", "qty": srt, "unit_price": rp, "total": round(srt*rp)},
    ]
    labor.append({"name": "Бетон цутгалт — хучилт", "unit": "м³", "qty": sv, "unit_price": cwp, "total": round(sv*cwp)})

    # 6. ШАТНЫ БҮТЭЦ
    if floors >= 2:
        stc = round(floors * 2.5, 1)
        str_ = round(floors * 0.2, 2)
        materials += [
            {"name": "Бетон зуурмаг М250 — шат", "unit": "м³", "qty": stc, "unit_price": cp, "total": round(stc*cp)},
            {"name": "Арматур — шат", "unit": "тонн", "qty": str_, "unit_price": rp, "total": round(str_*rp)},
        ]
        labor.append({"name": "Шатны бетон цутгалт", "unit": "м³", "qty": stc, "unit_price": cwp, "total": round(stc*cwp)})

    # 7. ДЭЭВЭР
    if "хавтгай" in roof_type.lower():
        rc = round(roof_area * 0.15, 1)
        rbp = get_material_price("mat_roof", "Рубероид") or 18000
        materials += [
            {"name": "Бетон — хавтгай дээвэр", "unit": "м³", "qty": rc, "unit_price": cp, "total": round(rc*cp)},
            {"name": "Рубероид — ус тусгаарлалт", "unit": "рулон", "qty": round(roof_area/10), "unit_price": rbp, "total": round(roof_area/10*rbp)},
        ]
        labor.append({"name": "Хавтгай дээвэр хийх", "unit": "м²", "qty": round(roof_area,1), "unit_price": 25000, "total": round(roof_area*25000)})
    else:
        if "металл черепица" in roof_type.lower() or "метал" in roof_type.lower():
            rmp = get_material_price("mat_roof", "Металл черепица") or 45000
            rmn = "Металл черепица"
        else:
            rmp = get_material_price("mat_roof", "Профнастил") or 21000
            rmn = "Профнастил"
        wdp = get_material_price("mat_wood", "Тавцан мод") or 5250
        wdq = round(roof_area * 12)
        rwp2 = get_material_price("labor_special", "Төмөр дээвэр") or 26500
        materials += [
            {"name": f"{rmn} — дээвэр", "unit": "м²", "qty": round(roof_area,1), "unit_price": rmp, "total": round(roof_area*rmp)},
            {"name": "Тавцан мод — дээврийн каркас", "unit": "м", "qty": wdq, "unit_price": wdp, "total": wdq*wdp},
        ]
        labor.append({"name": "Дээвэр угсралт", "unit": "м²", "qty": round(roof_area,1), "unit_price": rwp2, "total": round(roof_area*rwp2)})

    # 8. ДОТОР ЗАСАЛ — шал
    wa = round(total_area * 0.25, 1)
    da = round(total_area * 0.75, 1)
    tp = get_material_price("mat_interior", "Керамик плита") or 57000
    tw = get_material_price("labor_general", "Плита наалт") or 40000
    sw = get_material_price("labor_general", "Шалны тэгшилгээ") or 16500
    materials.append({"name": "Керамик плита — ванн, гал тогоо", "unit": "м²", "qty": wa, "unit_price": tp, "total": round(wa*tp)})
    labor.append({"name": "Плита тавих — ванн, гал тогоо", "unit": "м²", "qty": wa, "unit_price": tw, "total": round(wa*tw)})
    floor_mat = data.get("floor_material", "Ламинат")
    if "паркет" in floor_mat.lower():
        fp2 = get_material_price("mat_interior", "Паркет") or 140000
        materials.append({"name": "Паркет — үндсэн өрөө", "unit": "м²", "qty": da, "unit_price": fp2, "total": round(da*fp2)})
        labor.append({"name": "Паркет тавих", "unit": "м²", "qty": da, "unit_price": 26000, "total": round(da*26000)})
    else:
        lp = get_material_price("mat_interior", "Ламинат шал") or 42000
        lw = get_material_price("labor_general", "Ламинат шал тавих") or 19000
        materials.append({"name": "Ламинат шал — үндсэн өрөө", "unit": "м²", "qty": da, "unit_price": lp, "total": round(da*lp)})
        labor.append({"name": "Ламинат шал тавих", "unit": "м²", "qty": da, "unit_price": lw, "total": round(da*lw)})
    labor.append({"name": "Шалны стяжка", "unit": "м²", "qty": total_area, "unit_price": sw, "total": round(total_area*sw)})

    # 9. ДОТОР ХАНЫН ЗАСАЛ
    tis = round((net_wall_area + inner_wall_area * 2) * 0.85)
    zp = get_material_price("mat_interior", "Өнгөлгөөний цагаан замазка") or 12000
    zq = round(tis * 1.2)
    pip = get_material_price("mat_interior", "Дотор будаг") or 7500
    piq = round(tis * 0.3)
    materials += [
        {"name": "Цагаан замазка — дотор хана, тааз", "unit": "кг", "qty": zq, "unit_price": zp, "total": zq*zp},
        {"name": "Дотор эмульс будаг", "unit": "кг", "qty": piq, "unit_price": pip, "total": piq*pip},
    ]
    zwp = get_material_price("labor_special", "Цагаан замаска") or 14000
    pwp = get_material_price("labor_special", "Эмульс хийх") or 6750
    labor += [
        {"name": "Замазка хийх — дотор хана, тааз", "unit": "м²", "qty": tis, "unit_price": zwp, "total": tis*zwp},
        {"name": "Эмульс будаг — дотор хана, тааз", "unit": "м²", "qty": tis, "unit_price": pwp, "total": tis*pwp},
    ]

    # 10. ДОТОР ХУВААЛТ
    ibp = get_material_price("mat_brick", "Мак блок (20") or 7000
    iwp2 = get_material_price("labor_general", "Блокон хана өрөх /25") or 25000
    ibq = round(inner_wall_area * 16)
    materials.append({"name": "Мак блок (20см) — дотор хуваалт", "unit": "ш", "qty": ibq, "unit_price": ibp, "total": ibq*ibp})
    labor.append({"name": "Дотор хуваалт өрөх", "unit": "м²", "qty": round(inner_wall_area,1), "unit_price": iwp2, "total": round(inner_wall_area*iwp2)})

    # 11. ЦОНХ, ХААЛГА
    try:
        wna = int(str(data.get("windows","5")).split("-")[0])
    except:
        wna = 5
    try:
        dna = int(str(data.get("doors","4")).split("-")[0])
    except:
        dna = 4
    upf = int(str(data.get("units_per_floor",1)).replace("+",""))
    tu = max(1, upf * floors)
    tw2 = tu * wna
    td = tu * dna
    wnp = get_material_price("mat_window", "PVC цонх (1.2") or 450000
    dnp = get_material_price("mat_window", "Дотор хаалга") or 420000
    odp = get_material_price("mat_window", "Гадна хаалга (металл)") or 1150000
    wip = get_material_price("labor_special", "Хаалга угсрах (дотор") or 65000
    materials += [
        {"name": "PVC цонх (1.2x1.2м)", "unit": "ш", "qty": tw2, "unit_price": wnp, "total": tw2*wnp},
        {"name": "Дотор хаалга", "unit": "ш", "qty": td, "unit_price": dnp, "total": td*dnp},
        {"name": "Гадна хаалга (металл)", "unit": "ш", "qty": tu, "unit_price": odp, "total": tu*odp},
    ]
    labor.append({"name": "Цонх, хаалга угсралт", "unit": "ш", "qty": tw2+td, "unit_price": wip, "total": (tw2+td)*wip})

    # 12. ЦАХИЛГААНЫ СИСТЕМ
    cbp = get_material_price("mat_electrical", "Кабель ВВГ") or 3600
    cbq = round(total_area * 4)
    swp = get_material_price("mat_electrical", "Унтраалга 1") or 6100
    sop = get_material_price("mat_electrical", "Розетка 1") or 6600
    swq = round(tu * 8)
    soq = round(tu * 12)
    ewp = get_material_price("labor_general", "Цахилгаанчин") or 125000
    edp = round(total_area / 15)
    materials += [
        {"name": "Цахилгааны кабель ВВГ", "unit": "м", "qty": cbq, "unit_price": cbp, "total": cbq*cbp},
        {"name": "Унтраалга", "unit": "ш", "qty": swq, "unit_price": swp, "total": swq*swp},
        {"name": "Розетка", "unit": "ш", "qty": soq, "unit_price": sop, "total": soq*sop},
    ]
    labor.append({"name": "Цахилгааны ажил", "unit": "өдөр", "qty": edp, "unit_price": ewp, "total": edp*ewp})

    # 13. ИНЖЕНЕРИЙН СИСТЕМ
    htp = get_material_price("mat_plumbing", "Халаалтын систем") or 6750000
    wtp = get_material_price("mat_plumbing", "Цэвэр усны систем") or 1025000
    sgp = get_material_price("mat_plumbing", "Бохир усны систем") or 200000
    sap = get_material_price("mat_plumbing", "Ариун цэврийн өрөөний") or 1850000
    materials += [
        {"name": "Халаалтын систем", "unit": "багц", "qty": tu, "unit_price": htp, "total": tu*htp},
        {"name": "Цэвэр усны систем", "unit": "багц", "qty": tu, "unit_price": wtp, "total": tu*wtp},
        {"name": "Бохир усны систем", "unit": "багц", "qty": tu, "unit_price": sgp, "total": tu*sgp},
        {"name": "Ариун цэврийн тоноглол", "unit": "багц", "qty": tu, "unit_price": sap, "total": tu*sap},
    ]

    # 14. ТЭЭВЭР
    trp = get_material_price("transport_material", "Материал тээвэр") or 90000
    wsp = get_material_price("transport_material", "Хог зайлуулах") or 80000
    trt = max(10, round(total_area / 15))
    transport += [
        {"name": "Материал тээвэр", "unit": "удаа", "qty": trt, "unit_price": trp, "total": trt*trp},
        {"name": "Хог зайлуулах", "unit": "удаа", "qty": round(trt/2), "unit_price": wsp, "total": round(trt/2)*wsp},
    ]

    # 15. БУСАД
    duration = max(6, round(total_area / 60))
    dsp = get_material_price("other_design", "Архитектурын") or 25000
    pmp = get_material_price("other_permit", "Барилгын зөвшөөрөл") or 1250000
    inp2 = get_material_price("other_insurance", "Барилгын даатгал") or 1250000
    other += [
        {"name": "Архитектур, инженерийн зураг төсөл", "unit": "м²", "qty": round(total_area), "unit_price": dsp, "total": round(total_area*dsp)},
        {"name": "Барилгын зөвшөөрөл", "unit": "удаа", "qty": 1, "unit_price": pmp, "total": pmp},
        {"name": "Инженер хяналт", "unit": "сар", "qty": duration, "unit_price": 1750000, "total": duration*1750000},
        {"name": "Барилгын даатгал", "unit": "жил", "qty": max(1,round(duration/12)), "unit_price": inp2, "total": max(1,round(duration/12))*inp2},
    ]

    mt = round(sum(i["total"] for i in materials) * quality_coef)
    lt = round(sum(i["total"] for i in labor) * quality_coef)
    tt = sum(i["total"] for i in transport)
    ot = sum(i["total"] for i in other)
    gt = mt + lt + tt + ot
    ppm2 = round(gt / total_area) if total_area > 0 else 0

    return {
        "building_info": {
            "type": data.get("building_type",""),
            "area": f"{total_area:.0f} м²",
            "floors": str(floors),
            "location": data.get("location","Улаанбаатар"),
            "quality": data.get("quality","дунд"),
        },
        "materials": materials,
        "labor": labor,
        "transport": transport,
        "other": other,
        "summary": {
            "materials_total": mt,
            "labor_total": lt,
            "transport_total": tt,
            "other_total": ot,
            "grand_total": gt,
            "price_per_m2": ppm2,
            "duration_months": duration,
        },
        "notes": "⚠️ Энэ тооцоо ойролцоо үнэлгээ бөгөөд мэргэжлийн инженерийн тооцоог орлохгүй. Бодит зардал газрын байршил, ханган нийлүүлэгч, нарийн зураг төслөөс хамаарч өөрчлөгдөж болно. Гэрээ байгуулахаасаа өмнө мэргэжлийн байгууллагаар нарийвчилсан тооцоо гаргуулна уу."
    }

def budget_calculator(request):
    from django.conf import settings
    result = None
    error = None
    
    if request.method == "POST":
        building_type = request.POST.get("building_type", "")
        floors = request.POST.get("floors", "1")
        quality = request.POST.get("quality", "дунд")
        location = request.POST.get("location", "Улаанбаатар")
        build_year = request.POST.get("build_year", "2026")
        length = request.POST.get("length", "")
        width = request.POST.get("width", "")
        total_height = request.POST.get("total_height", "")
        ceiling_height = request.POST.get("ceiling_height", "2.7")
        inner_wall_length = request.POST.get("inner_wall_length", "")
        windows = request.POST.get("windows", "")
        doors = request.POST.get("doors", "")
        units_per_floor = request.POST.get("units_per_floor", "4")
        ground_floor_units = request.POST.get("ground_floor_units", "Дээрх давхартай адил")
        # ШИНЭ: Python норм тооцоо — AI ашиглахгүй
        if length and width:
            try:
                norm_data = {
                    "building_type": building_type,
                    "floors": floors,
                    "length": length,
                    "width": width,
                    "ceiling_height": ceiling_height,
                    "wall_material": request.POST.get("wall_material", "Мак блок"),
                    "insulation": request.POST.get("insulation", ""),
                    "foundation_type": request.POST.get("foundation_type", "Шугаман суурь"),
                    "foundation_depth": request.POST.get("foundation_depth", "2.5"),
                    "roof_type": request.POST.get("roof_type", ""),
                    "floor_material": request.POST.get("floor_material", "Ламинат"),
                    "facade": request.POST.get("facade", "Шавар штукатур"),
                    "wall_finish": request.POST.get("wall_finish", "Хосолсон"),
                    "electrical": request.POST.get("electrical", "Стандарт 220В"),
                    "heating": request.POST.get("heating", ""),
                    "windows": windows,
                    "doors": doors,
                    "units_per_floor": units_per_floor,
                    "quality": quality,
                    "location": location,
                }
                result = calculate_budget_norm(norm_data)
            except Exception as e:
                error = f"Тооцооны алдаа: {str(e)}"

        # Нийт айлын тоо тооцоолох
        try:
            fl = int(floors) if floors else 1
            upf = int(units_per_floor.replace("+","")) if units_per_floor else 4
            if ground_floor_units == "Дээрх давхартай адил":
                total_units = upf * fl
            elif "Хагас" in ground_floor_units:
                total_units = (upf // 2) + upf * (fl - 1)
            elif "Бүгд нийтийн" in ground_floor_units or "Гараж" in ground_floor_units:
                total_units = upf * (fl - 1)
            else:
                total_units = upf * fl
            total_units_text = f"{total_units} айл ({upf} айл × {fl} давхар)"
        except:
            total_units = 0
            total_units_text = "тодорхойгүй" 
        foundation_type = request.POST.get("foundation_type", "")
        foundation_depth = request.POST.get("foundation_depth", "")
        foundation_width = request.POST.get("foundation_width", "")
        concrete_grade = request.POST.get("concrete_grade", "М250")
        soil_type = request.POST.get("soil_type", "")
        water_table = request.POST.get("water_table", "")
        wall_material = request.POST.get("wall_material", "")
        wall_thickness = request.POST.get("wall_thickness", "")
        insulation = request.POST.get("insulation", "")
        inner_wall_material = request.POST.get("inner_wall_material", "")
        roof_type = request.POST.get("roof_type", "")
        facade = request.POST.get("facade", "")
        floor_material = request.POST.get("floor_material", "")
        wall_finish = request.POST.get("wall_finish", "")
        heating = request.POST.get("heating", "")
        water = request.POST.get("water", "")
        electrical = request.POST.get("electrical", "")
        extras = request.POST.get("extras", "")
        
    import json as _json
    result_json = _json.dumps(result, ensure_ascii=False) if result else "{}"
    return render(request, "registry/budget_calculator.html", {
        "result": result,
        "result_json": result_json,
        "error": error,
        "post_data": request.POST,
        "display_name": get_display_name(request.user),
    })


def budget_excel(request):
    import json
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    data_str = request.POST.get("data", request.GET.get("data", "{}"))
    try:
        data = json.loads(data_str)
    except:
        return HttpResponse("Алдаа", status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Төсөв"

    # Styles
    header_fill = PatternFill("solid", fgColor="1e3a4a")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill("solid", fgColor="f59e0b")
    section_font = Font(bold=True, color="1e3a4a", size=11)
    total_fill = PatternFill("solid", fgColor="fef3c7")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    row = 1

    # Title
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "БАРИЛГЫН ТӨСВИЙН ТООЦОО"
    ws[f"A{row}"].font = Font(bold=True, size=14, color="1e3a4a")
    ws[f"A{row}"].alignment = center
    row += 1

    # Building info
    info = data.get("building_info", {})
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"{info.get('type','')} | {info.get('area','')} | {info.get('location','')} | {info.get('quality','')}"
    ws[f"A{row}"].font = Font(italic=True, color="64748b")
    ws[f"A{row}"].alignment = center
    row += 2

    def write_section(title, items, total_key):
        nonlocal row
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        headers = ["Нэр", "Нэгж", "Тоо хэмжээ", "Нэгж үнэ (₮)", "Нийт (₮)"]
        aligns = ["left", "center", "right", "right", "right"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal=aligns[i], vertical="center")
            cell.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        # Data rows
        for item in items:
            ws.cell(row=row, column=1, value=item.get("name","")).border = border
            ws.cell(row=row, column=2, value=item.get("unit","")).alignment = center
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=3, value=item.get("qty",0)).alignment = right
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4, value=item.get("unit_price",0)).number_format = "#,##0"
            ws.cell(row=row, column=4).alignment = right
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=5, value=item.get("total",0)).number_format = "#,##0"
            ws.cell(row=row, column=5).alignment = right
            ws.cell(row=row, column=5).border = border
            row += 1

        # Total row
        summary = data.get("summary", {})
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Нийт дүн"
        ws[f"A{row}"].fill = total_fill
        ws[f"A{row}"].font = total_font
        ws[f"A{row}"].border = border
        ws[f"E{row}"] = summary.get(total_key, 0)
        ws[f"E{row}"].number_format = "#,##0"
        ws[f"E{row}"].fill = total_fill
        ws[f"E{row}"].font = total_font
        ws[f"E{row}"].alignment = right
        ws[f"E{row}"].border = border
        row += 2

    write_section("🧱 МАТЕРИАЛЫН ЗАРДАЛ", data.get("materials", []), "materials_total")
    write_section("👷 АЖИЛЧДЫН ЗАРДАЛ", data.get("labor", []), "labor_total")
    write_section("📦 БУСАД ЗАРДАЛ", data.get("other", []), "other_total")

    # Grand total
    summary = data.get("summary", {})
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "НИЙТ ТӨСӨВ"
    ws[f"A{row}"].fill = PatternFill("solid", fgColor="f59e0b")
    ws[f"A{row}"].font = Font(bold=True, size=13, color="1e3a4a")
    ws[f"A{row}"].border = border
    ws[f"E{row}"] = summary.get("grand_total", 0)
    ws[f"E{row}"].number_format = "#,##0"
    ws[f"E{row}"].fill = PatternFill("solid", fgColor="f59e0b")
    ws[f"E{row}"].font = Font(bold=True, size=13, color="1e3a4a")
    ws[f"E{row}"].alignment = right
    ws[f"E{row}"].border = border
    ws.row_dimensions[row].height = 26
    row += 1

    # Per m2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "1 м² үнэ"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"E{row}"] = summary.get("price_per_m2", 0)
    ws[f"E{row}"].number_format = "#,##0"
    ws[f"E{row}"].alignment = right
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Барилгын хугацаа"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"E{row}"] = f"{summary.get('duration_months', 0)} сар"
    ws[f"E{row}"].alignment = right
    row += 2

    # Notes
    if data.get("notes"):
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = f"Анхаарах зүйлс: {data['notes']}"
        ws[f"A{row}"].font = Font(italic=True, color="166534")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="f0fdf4")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=budget.xlsx"
    wb.save(response)
    return response


def budget_chat(request):
    from django.conf import settings
    from django.http import JsonResponse
    import anthropic, json

    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        body = json.loads(request.body)
        messages = body.get("messages", [])
        user_msg = body.get("message", "")
    except:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    if not user_msg:
        return JsonResponse({"error": "Хоосон асуулт"}, status=400)

    system_prompt = """Та Монголын барилгын салбарын туршлагатай мэргэжилтэн. 
Барилга барихаар төлөвлөж байгаа хүмүүст практик зөвлөгөө өгнө.

Та дараах мэдлэгтэй:
- Монголын барилгын норм, дүрэм (БНбД)
- Барилгын материал, технологи
- Зөвшөөрөл, бүртгэлийн процесс
- Зураг төсөл, инженерийн шийдэл
- Барилгын компани сонгох зөвлөгөө
- Барилгын хугацаа, зардлын тооцоо
- Монголын цаг уур, газар хөрсний онцлог

Хариултаа товч, практик, Монгол хэлээр өгнө үү. 
Хэрэв тооцоо хийх шаардлагатай бол /budget/ хуудсыг ашиглахыг санал болгоно уу."""

    try:
        client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
        
        # Өмнөх яриаг нэмэх
        api_messages = []
        for msg in messages[-10:]:  # Сүүлийн 10 мессеж
            api_messages.append({
                "role": msg["role"],
                "content": msg["content"]
            })
        api_messages.append({"role": "user", "content": user_msg})

        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            system=system_prompt,
            messages=api_messages
        )
        reply = response.content[0].text
        return JsonResponse({"reply": reply})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def budget_file_upload(request):
    from django.conf import settings
    result = None
    error = None

    if request.method == "POST":
        import anthropic, openpyxl, json
        from io import BytesIO

        uploaded_file = request.FILES.get("budget_file")
        if not uploaded_file:
            error = "Файл оруулаагүй байна."
        else:
            try:
                # Excel файл унших
                wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()))
                ws = wb.active

                rows_data = []
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows_data.append([str(c) if c is not None else "" for c in row])

                # DB-аас үнэ авах
                from apps.public.models import MaterialPrice
                key_prices = [
                    ("mat_cement", "Цемент"),
                    ("mat_sand", "Элс, хайрга"),
                    ("mat_brick", "Тоосго, блок"),
                    ("mat_rebar", "Арматур"),
                    ("mat_wood", "Мод"),
                    ("mat_insulation", "Дулаалга"),
                    ("mat_window", "Цонх, хаалга"),
                    ("mat_interior", "Дотор засал"),
                    ("mat_plumbing", "Сантехник"),
                    ("mat_electrical", "Цахилгаан"),
                    ("labor_general", "Барилгачин"),
                    ("labor_special", "Мэргэжилтэн"),
                    ("transport_material", "Тээвэр"),
                    ("machine_crane", "Кран"),
                    ("machine_excavator", "Экскаватор"),
                ]
                price_lines = []
                for cat, label in key_prices:
                    items = MaterialPrice.objects.filter(is_active=True, category=cat)[:2]
                    for p in items:
                        price_lines.append(f"- {p.name}: {int(p.price_min):,}₮-{int(p.price_max):,}₮/{p.unit}")
                price_text = "\n".join(price_lines[:40])

                # Excel өгөгдлийг текст болгох
                excel_text = "\n".join([" | ".join(row) for row in rows_data[:80]])

                prompt = f"""Та Монголын барилгын салбарын мэргэжилтэн. Дараах Excel файлын өгөгдлөөс барилгын төсвийг тооцоолж өгнө үү.

Excel файлын агуулга:
{excel_text}

Одоогийн зах зээлийн үнэ:
{price_text}

Excel дээр тоо хэмжээ оруулсан бол тэр тоог ашиглана уу.
Тоо хэмжээ байхгүй бол хоосон орхино уу.
Нэгж үнэ байхгүй бол дээрх зах зээлийн үнийг ашиглана уу.

ЗӨВХӨН JSON форматаар хариу өгнө үү:
{{
  "building_info": {{"type": "файлаас авсан барилгын нэр", "area": "талбай", "location": "", "quality": ""}},
  "materials": [{{"name": "нэр", "unit": "нэгж", "qty": тоо, "unit_price": үнэ, "total": нийт}}],
  "labor": [{{"name": "нэр", "unit": "нэгж", "qty": тоо, "unit_price": үнэ, "total": нийт}}],
  "transport": [{{"name": "нэр", "unit": "нэгж", "qty": тоо, "unit_price": үнэ, "total": нийт}}],
  "other": [{{"name": "нэр", "unit": "нэгж", "qty": тоо, "unit_price": үнэ, "total": нийт}}],
  "summary": {{"materials_total": 0, "labor_total": 0, "transport_total": 0, "other_total": 0, "grand_total": 0, "price_per_m2": 0, "duration_months": 0}},
  "notes": "тайлбар"
}}"""

                client = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
                message = client.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=4000,
                    messages=[{"role": "user", "content": prompt}]
                )
                import re
                raw = message.content[0].text.strip()
                raw = re.sub(r"^```json\s*", "", raw)
                raw = re.sub(r"^```\s*", "", raw)
                raw = re.sub(r"\s*```$", "", raw)
                raw = raw.strip()
                start = raw.find("{")
                end = raw.rfind("}") + 1
                if start >= 0 and end > start:
                    raw = raw[start:end]
                result = json.loads(raw)

            except Exception as e:
                error = f"Алдаа: {str(e)}"

    import json as _json
    result_json = _json.dumps(result, ensure_ascii=False) if result else "{}"
    return render(request, "registry/budget_file.html", {
        "result": result,
        "result_json": result_json,
        "error": error,
        "display_name": get_display_name(request.user),
    })
