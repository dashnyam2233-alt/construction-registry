from __future__ import annotations

from io import BytesIO
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.core.models import (
    CITY_CHOICES,
    UB_DISTRICT_CHOICES,
    COMPANY_ACTIVITY_DIRECTION_CHOICES,
    RESPONSIBLE_ROLE_CHOICES,
    ENGINEER_SPECIALTY_CHOICES,
)


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 70)


def _add_choices_sheet(wb: Workbook):
    ws = wb.create_sheet("choices (сонголтууд)")

    sections: List[Tuple[str, List[Tuple[str, str]]]] = [
        ("Хот/Аймаг (CITY_CHOICES)", CITY_CHOICES),
        ("УБ Дүүрэг (UB_DISTRICT_CHOICES)", UB_DISTRICT_CHOICES),
        ("Компани - Үйл ажиллагааны чиглэл (COMPANY_ACTIVITY_DIRECTION_CHOICES)", COMPANY_ACTIVITY_DIRECTION_CHOICES),
        ("Ажилтан - Хариуцсан ажил (RESPONSIBLE_ROLE_CHOICES)", RESPONSIBLE_ROLE_CHOICES),
        ("Ажилтан - Инженерийн төрөл (ENGINEER_SPECIALTY_CHOICES)", ENGINEER_SPECIALTY_CHOICES),
        ("Ажилтан - Хүйс (Gender)", [("male", "Эр"), ("female", "Эм")]),
        ("Ажилтан - Мэргэжил (Profession)", [
            ("engineer", "Инженер"),
            ("architect", "Архитектор"),
            ("foreman", "Даамал"),
            ("accountant", "Нягтлан"),
            ("hr", "Хүний нөөц"),
            ("manager", "Менежер"),
            ("worker", "Ажилчин"),
            ("other", "Бусад"),
        ]),
    ]

    r = 1
    ws.cell(r, 1).value = "Сонголт"
    ws.cell(r, 2).value = "code"
    ws.cell(r, 3).value = "label (Монгол)"
    r += 1

    for title, items in sections:
        ws.cell(r, 1).value = title
        r += 1
        for code, label in items:
            ws.cell(r, 2).value = code
            ws.cell(r, 3).value = label
            r += 1
        r += 2

    _autosize(ws)


def _add_readme_company(wb: Workbook):
    ws = wb.create_sheet("README (Заавар)")

    lines = [
        "КОМПАНИ IMPORT — ЗААВАР",
        "",
        "1) Энэ файлын 'company_import' sheet дээр мэдээллээ бөглөнө.",
        "2) Сонголтот талбарууд дээр та 2 янзаар бөглөж болно:",
        "   - code (ж: UB, CONSTRUCTION, BZD гэх мэт)",
        "   - Монгол нэр (ж: Улаанбаатар, Барилга угсралт, Баянзүрх гэх мэт)",
        "3) 'РД/Бүртгэлийн №' нь компанийн гол түлхүүр. Давхардвал тухайн компани UPDATE хийгдэнэ.",
        "4) 'Дүүрэг (УБ үед)' талбар зөвхөн Хот/Аймаг = UB үед ашиглагдана. Бусад үед хоосон байж болно.",
        "5) Import унах шалтгаанууд:",
        "   - Үйл ажиллагааны чиглэл / Хот/Аймаг / Дүүрэг сонголт буруу байх",
        "   - Формат эвдэрсэн байх (ж: нүдээ олсон багана нэрүүдийг өөрчлөх)",
        "",
        "Зөвлөгөө:",
        "- Сонголтуудыг 'choices (сонголтууд)' sheet-ээс copy хийж бөглөөрэй.",
    ]
    for i, t in enumerate(lines, start=1):
        ws.cell(i, 1).value = t
    ws.column_dimensions["A"].width = 110


def _add_readme_worker(wb: Workbook):
    ws = wb.create_sheet("README (Заавар)")

    lines = [
        "АЖИЛТАН IMPORT — ЗААВАР",
        "",
        "1) Энэ файлын 'worker_import' sheet дээр мэдээллээ бөглөнө.",
        "2) 'Компани РД/Бүртгэлийн №' заавал зөв байх ёстой. Олдохгүй бол тухайн мөр ERROR болж import унадаг.",
        "3) Сонголтот талбарууд дээр та 2 янзаар бөглөж болно:",
        "   - code (ж: male, engineer, ENGINEER, UB, BZD гэх мэт)",
        "   - Монгол нэр (ж: Эр, Инженер, Инженер, Улаанбаатар, Баянзүрх гэх мэт)",
        "4) 'Регистр' нь ажилтны түлхүүр. Давхардвал тухайн ажилтан UPDATE хийгдэнэ.",
        "5) 'Хариуцсан ажил' = ENGINEER (Инженер) бол 'Инженерийн төрөл' заавал бөглөнө.",
        "6) 'Гэрлэсэн эсэх' дээр зөвхөн: Тийм / Үгүй (эсвэл 1/0, True/False) ашиглана.",
        "7) Import унах шалтгаанууд:",
        "   - Сонголт буруу (Хүйс, Мэргэжил, Хот/Аймаг гэх мэт)",
        "   - Компани олдохгүй (FK error)",
        "   - ENGINEER үед Инженерийн төрөл хоосон",
        "",
        "Зөвлөгөө:",
        "- Сонголтуудыг 'choices (сонголтууд)' sheet-ээс copy хийж бөглөөрэй.",
        "- Эхлээд Company import хийгээд, дараа нь Worker import хийх нь зөв.",
    ]
    for i, t in enumerate(lines, start=1):
        ws.cell(i, 1).value = t
    ws.column_dimensions["A"].width = 110


def _add_readme_simple(
    wb: Workbook,
    title: str,
    sheet_name: str,
    extra_lines: List[str] | None = None,
):
    ws = wb.create_sheet("README (Заавар)")
    lines = [
        f"{title} — ЗААВАР",
        "",
        f"1) Энэ файлын '{sheet_name}' sheet дээр мэдээллөө бөглөнө.",
        "2) Column header-уудын нэрийг өөрчлөхгүй (өөрчилбөл import унах эрсдэлтэй).",
        "3) Эхлээд 1–2 мөрөөр туршиж import хийгээд дараа нь бөөнөөр нь import хийгээрэй.",
    ]
    if extra_lines:
        lines += [""] + extra_lines

    for i, t in enumerate(lines, start=1):
        ws.cell(i, 1).value = t
    ws.column_dimensions["A"].width = 110


def build_company_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "company_import"

    headers = [
        "Компани нэр",
        "РД/Бүртгэлийн №",
        "Үйл ажиллагааны чиглэл",
        "Дэд сонголт",
        "Хот/Аймаг",
        "Дүүрэг (УБ үед)",
        "Дэлгэрэнгүй хаяг",
        "Утас",
        "Имэйл",
        "Вэб",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "Жишээ компани ХХК",
        "AB12345678",
        "CONSTRUCTION",
        "Барилга угсралтын үйл ажиллагаа",
        "UB",
        "BZD",
        "Дэлгэрэнгүй хаяг ...",
        "99112233",
        "info@example.com",
        "example.com",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_company(wb)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_worker_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "worker_import"

    headers = [
        "Ургийн овог",
        "Эцэг/эхийн нэр",
        "Нэр",
        "Хүйс",
        "Регистр",
        "Төрсөн огноо",
        "Төрсөн газар - Аймаг/Хот",
        "Төрсөн газар - Сум/Дүүрэг",
        "Гэрлэсэн эсэх",
        "Мэргэжил",
        "Компани РД/Бүртгэлийн №",
        "Хариуцсан ажил",
        "Инженерийн төрөл",
        "Утас",
        "Имэйл",
        "Facebook хаяг",
        "Instagram хаяг",
        "Viber хаяг/дугаар",
        "Хот/Аймаг",
        "Дүүрэг (УБ үед)",
        "Оршин суугаа газрын хаяг",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "Бат",
        "Дорж",
        "Төгсөө",
        "Эр",
        "УБ12345678",
        "1990-01-01",
        "UB",
        "BZD",
        "Үгүй",
        "Инженер",
        "AB12345678",
        "Инженер",
        "Иргэний ба үйлдвэрлэлийн барилгын инженер",
        "99001122",
        "tugsuu@example.com",
        "",
        "",
        "",
        "UB",
        "BZD",
        "Дэлгэрэнгүй хаяг",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_worker(wb)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -------------------------------------------------------------------
# ✅ ШИНЭ MN IMPORT ЗАГВАРУУД
# -------------------------------------------------------------------

def build_family_member_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "family_member_import"

    headers = [
        "Ажилтны регистр (заавал)",
        "Хамаарал",
        "Ургийн овог",
        "Эцэг/эхийн нэр",
        "Нэр",
        "Хүйс",
        "Регистр",
        "Төрсөн огноо",
        "Утас",
        "Имэйл",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "УБ12345678",
        "Эхнэр/Нөхөр",
        "",
        "",
        "Жишээ",
        "",
        "",
        "1992-01-01",
        "99112233",
        "",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_simple(
        wb,
        title="АЖИЛТНЫ ХАМААРАЛ IMPORT",
        sheet_name="family_member_import",
        extra_lines=[
            "- Эхлээд Worker import хийгдсэн байх хэрэгтэй.",
            "- 'Ажилтны регистр' буруу бол FK алдаа гарч import унах магадлалтай.",
        ],
    )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_brigade_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "brigade_import"

    headers = [
        "Бригадын нэр",
        "Үйл ажиллагааны чиглэл",
        "Харьяалах компани РД/Бүртгэлийн №",
        "Ахлагч ажилтны регистр",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "Жишээ бригад",
        "Барилга угсралт",
        "AB12345678",
        "УБ12345678",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_simple(
        wb,
        title="БАРИЛГЫН БРИГАД IMPORT",
        sheet_name="brigade_import",
        extra_lines=[
            "- Эхлээд Company + Worker бүртгэсэн байх хэрэгтэй.",
        ],
    )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_brigade_member_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "brigade_member_import"

    headers = [
        "Бригадын нэр (эсвэл ID)",
        "Ажилтны регистр (заавал)",
        "Гишүүний үүрэг/албан тушаал",
        "Эхэлсэн огноо",
        "Дууссан огноо",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "Жишээ бригад",
        "УБ12345678",
        "Гишүүн",
        "2025-01-01",
        "",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_simple(
        wb,
        title="БРИГАДЫН ГИШҮҮН IMPORT",
        sheet_name="brigade_member_import",
        extra_lines=[
            "- Эхлээд Brigade + Worker бүртгэсэн байх хэрэгтэй.",
        ],
    )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_government_org_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "government_org_import"

    headers = [
        "Байгууллагын нэр",
        "РД/Бүртгэлийн №",
        "Хот/Аймаг",
        "Дүүрэг (УБ үед)",
        "Хаяг",
        "Утас",
        "Имэйл",
        "Вэб",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "Жишээ төрийн байгууллага",
        "GA12345678",
        "UB",
        "BZD",
        "Хаяг ...",
        "99112233",
        "",
        "",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_simple(
        wb,
        title="САЛБАРЫН ТӨРИЙН БАЙГУУЛЛАГА IMPORT",
        sheet_name="government_org_import",
    )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_non_government_org_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ngo_import"

    headers = [
        "Байгууллагын нэр",
        "РД/Бүртгэлийн №",
        "Хот/Аймаг",
        "Дүүрэг (УБ үед)",
        "Хаяг",
        "Утас",
        "Имэйл",
        "Вэб",
        "Тайлбар",
    ]
    ws.append(headers)

    ws.append([
        "Жишээ ТББ",
        "NG12345678",
        "UB",
        "BZD",
        "Хаяг ...",
        "99112233",
        "",
        "",
        "",
    ])

    _autosize(ws)
    _add_choices_sheet(wb)
    _add_readme_simple(
        wb,
        title="САЛБАРЫН ТӨРИЙН БУС БАЙГУУЛЛАГА IMPORT",
        sheet_name="ngo_import",
    )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
