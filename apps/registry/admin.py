# -*- coding: utf-8 -*-
from io import BytesIO
from datetime import datetime

from django.template.response import TemplateResponse
from django.urls import reverse, path
from django.contrib import admin, messages
from django import forms
from django.core.mail import send_mail
from django.db.models import Count, Q
from django.utils.html import format_html
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin, GroupAdmin as DjangoGroupAdmin
from django.db import models

from import_export.admin import ImportExportModelAdmin
from openpyxl import Workbook

from .import_resources import CompanyResource, WorkerResource
from .import_templates import (
    build_company_template_xlsx,
    build_worker_template_xlsx,
    build_family_member_template_xlsx,
    build_brigade_template_xlsx,
    build_brigade_member_template_xlsx,
    build_government_org_template_xlsx,
    build_non_government_org_template_xlsx,
)

from .models import (
    GovernmentOrganization, NonGovernmentOrganization, Company, Worker,
    FamilyMember, AdminGroup, Brigade, BrigadeMember, UserCompanyProfile,
    CITY_CHOICES, UB_DISTRICT_CHOICES, COMPANY_ACTIVITY_DIRECTION_CHOICES,
    normalize_search_text,
    SiteConfig, MessageLog,
)
from apps.public.models import Banner, PublicPost, HeroBanner, SliderAd, SubBanner
from .birth_soums import BIRTH_SOUMS

ROLE_ADMIN = "ADMIN_FULL"
ROLE_OPERATOR = "COMPANY_OPERATOR"
ROLE_VIEWER = "VIEWER"


def _is_admin_user(user) -> bool:
    if not user or not user.is_authenticated: return False
    return bool(user.is_superuser) or user.groups.filter(name=ROLE_ADMIN).exists()

def _is_operator(user) -> bool:
    if not user or not user.is_authenticated: return False
    return user.groups.filter(name=ROLE_OPERATOR).exists()

def _is_viewer(user) -> bool:
    if not user or not user.is_authenticated: return False
    return user.groups.filter(name=ROLE_VIEWER).exists()

def _user_company(user):
    if not user or not user.is_authenticated: return None
    prof = getattr(user, "company_profile", None)
    if not prof: return None
    return getattr(prof, "company", None)


VIEWER_ALLOWED_MODELS = {
    "company": True, "worker": True, "familymember": True,
    "brigade": True, "brigademember": True,
    "governmentorganization": True, "nongovernmentorganization": True,
}


def send_email_action(modeladmin, request, queryset):
    to_emails = [e for e in queryset.values_list("email", flat=True) if e]
    if not to_emails:
        messages.warning(request, "Имэйл хаягтай мөр сонгоогүй байна.")
        return
    send_mail("Мэдээлэл", "Танд мэдээлэл хүргэж байна.", None, to_emails, fail_silently=False)
    messages.success(request, f"Имэйл илгээлээ. ({len(to_emails)})")

send_email_action.short_description = "Имэйл илгээх (console)"

TEXT_FIELD_TYPES = {"CharField", "TextField", "EmailField", "URLField", "SlugField"}

def contains_any(field_lookup, term):
    t = (term or "").strip()
    if not t: return Q()
    return Q(**{f"{field_lookup}__icontains": t})

def model_has_field(model, field_name):
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False

def global_search_q(model, term):
    term = (term or "").strip()
    if not term: return Q()
    q = Q()
    if model_has_field(model, "search_normalized"):
        norm = normalize_search_text(term)
        if norm:
            q |= Q(search_normalized__contains=norm)
    for f in model._meta.get_fields():
        if not getattr(f, "concrete", False): continue
        if getattr(f, "is_relation", False): continue
        try:
            internal = f.get_internal_type()
        except Exception:
            continue
        if internal in TEXT_FIELD_TYPES:
            q |= contains_any(f.name, term)
    field_names = {ff.name for ff in model._meta.fields}
    if "company" in field_names:
        q |= contains_any("company__name", term)
    if "worker" in field_names:
        q |= contains_any("worker__first_name", term) | contains_any("worker__last_name", term)
    return q


class GlobalContainsSearchAdmin(admin.ModelAdmin):
    search_fields = ("id",)

    def get_search_results(self, request, queryset, search_term):
        term = (search_term or "").strip()
        if not term: return queryset, False
        q = global_search_q(queryset.model, term)
        if q == Q(): return queryset.none(), False
        return queryset.filter(q).distinct(), True


def build_xlsx_from_headers(headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(list(headers))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def default_model_headers(model):
    headers = []
    for f in model._meta.fields:
        if getattr(f, "primary_key", False): continue
        if isinstance(f, (models.AutoField, models.BigAutoField)): continue
        headers.append(f.name)
    return headers

def _fmt_cell_value(v):
    if v is None: return ""
    try:
        import datetime as _dt
        if isinstance(v, (_dt.date, _dt.datetime)): return v.isoformat()
    except Exception: pass
    if isinstance(v, bool): return "Тийм" if v else "Үгүй"
    return str(v)

def _mn_headers_and_getters(model, field_names):
    headers = []
    getters = []
    for name in field_names:
        try:
            f = model._meta.get_field(name)
            header = getattr(f, "verbose_name", name)
        except Exception:
            f = None
            header = name
        headers.append(str(header))
        if f is not None and getattr(f, "choices", None):
            def _make_choice_getter(nm):
                def g(obj):
                    try:
                        fn = getattr(obj, f"get_{nm}_display", None)
                        if callable(fn): return _fmt_cell_value(fn())
                    except Exception: pass
                    return _fmt_cell_value(getattr(obj, nm, ""))
                return g
            getters.append(_make_choice_getter(name))
        elif f is not None and getattr(f, "is_relation", False):
            def _make_rel_getter(nm):
                def g(obj):
                    try:
                        rel = getattr(obj, nm, None)
                        return _fmt_cell_value(rel) if rel else ""
                    except Exception: return ""
                return g
            getters.append(_make_rel_getter(name))
        else:
            def _make_plain_getter(nm):
                def g(obj):
                    try: return _fmt_cell_value(getattr(obj, nm, ""))
                    except Exception: return ""
                return g
            getters.append(_make_plain_getter(name))
    return headers, getters

def _write_sheet(wb, title_mn, headers, rows):
    ws = wb.create_sheet(title_mn)
    ws.append(list(headers))
    for r in rows: ws.append(list(r))

def _export_workbook_response(wb, filename):
    try:
        if wb.worksheets and wb.worksheets[0].title == "Sheet" and wb.worksheets[0].max_row == 1 and wb.worksheets[0].max_column == 1:
            wb.remove(wb.worksheets[0])
    except Exception: pass
    bio = BytesIO()
    wb.save(bio)
    resp = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def export_selected_companies_mn_xlsx(modeladmin, request, queryset):
    if not _is_admin_user(request.user):
        messages.error(request, "Зөвхөн админ экспорт хийх эрхтэй.")
        return None
    companies = list(queryset)
    if not companies:
        messages.warning(request, "Экспорт хийх компани сонгоогүй байна.")
        return None
    company_ids = [c.id for c in companies]
    workers = list(Worker.objects.filter(company_id__in=company_ids).select_related("company"))
    worker_ids = [w.id for w in workers]
    families = list(FamilyMember.objects.filter(worker_id__in=worker_ids).select_related("worker"))
    brigades = list(Brigade.objects.filter(companies__in=company_ids).distinct().prefetch_related("companies").select_related("leader_worker"))
    brigade_ids = [b.id for b in brigades]
    brigade_members = list(BrigadeMember.objects.filter(Q(brigade_id__in=brigade_ids) | Q(worker_id__in=worker_ids)).select_related("brigade", "worker").distinct())
    wb = Workbook()
    company_fields = ["name", "register_no", "activity_type", "activity_direction", "activity_sub_direction", "city", "district", "address", "phone", "email", "website", "note"]
    comp_headers, comp_getters = _mn_headers_and_getters(Company, company_fields)
    _write_sheet(wb, "Компани", comp_headers, [[g(c) for g in comp_getters] for c in companies])
    worker_fields = ["last_name", "parent_name", "first_name", "gender", "register_no", "birth_date", "birth_place_city", "birth_place_sub", "married", "profession", "company", "responsible_role", "engineer_specialty", "phone", "email", "facebook_url", "instagram_url", "viber", "city", "district", "address", "note"]
    w_headers, w_getters = _mn_headers_and_getters(Worker, worker_fields)
    _write_sheet(wb, "Ажиллагсад", w_headers, [[g(w) for g in w_getters] for w in workers])
    fam_fields = ["worker", "relation_type", "last_name", "first_name", "register_no", "birth_date", "phone", "email", "facebook_url", "instagram_url", "viber", "note"]
    f_headers, f_getters = _mn_headers_and_getters(FamilyMember, fam_fields)
    _write_sheet(wb, "Хамаарал", f_headers, [[g(fm) for g in f_getters] for fm in families])
    brigade_headers = ["Бригадын нэр", "Үйл ажиллагааны чиглэл", "Дэд сонголт(ууд)", "Ахлагч", "Хамтарч ажилладаг компаниуд", "Тайлбар"]
    b_rows = []
    for b in brigades:
        try: companies_txt = ", ".join([str(c) for c in b.companies.all()])
        except Exception: companies_txt = ""
        b_rows.append([_fmt_cell_value(getattr(b, "name", "")), _fmt_cell_value(b.get_activity_directions_display() if hasattr(b, "get_activity_directions_display") else getattr(b, "activity_directions_csv", "")), _fmt_cell_value(getattr(b, "activity_sub_directions_csv", "")), _fmt_cell_value(getattr(b, "leader_worker", "")) if getattr(b, "leader_worker_id", None) else "", companies_txt, _fmt_cell_value(getattr(b, "note", ""))])
    _write_sheet(wb, "Бригад", brigade_headers, b_rows)
    _write_sheet(wb, "Бригадын гишүүд", ["Бригад", "Бригадын гишүүн (Ажилтан)"], [[_fmt_cell_value(x.brigade), _fmt_cell_value(x.worker)] for x in brigade_members])
    return _export_workbook_response(wb, f"companies_export_mn_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

export_selected_companies_mn_xlsx.short_description = "Export (MN XLSX) — Сонгогдсон компани + ажиллагсад + бусад"

def export_selected_workers_mn_xlsx(modeladmin, request, queryset):
    if not _is_admin_user(request.user):
        messages.error(request, "Зөвхөн админ экспорт хийх эрхтэй.")
        return None
    workers = list(queryset.select_related("company"))
    if not workers:
        messages.warning(request, "Экспорт хийх ажилтан сонгоогүй байна.")
        return None
    worker_ids = [w.id for w in workers]
    families = list(FamilyMember.objects.filter(worker_id__in=worker_ids).select_related("worker"))
    company_ids = sorted({w.company_id for w in workers if w.company_id})
    companies = list(Company.objects.filter(id__in=company_ids)) if company_ids else []
    brigades = list(Brigade.objects.filter(Q(leader_worker_id__in=worker_ids) | Q(members__worker_id__in=worker_ids)).distinct().prefetch_related("companies").select_related("leader_worker"))
    brigade_ids = [b.id for b in brigades]
    brigade_members = list(BrigadeMember.objects.filter(Q(worker_id__in=worker_ids) | Q(brigade_id__in=brigade_ids)).distinct().select_related("brigade", "worker"))
    wb = Workbook()
    worker_fields = ["last_name", "parent_name", "first_name", "gender", "register_no", "birth_date", "birth_place_city", "birth_place_sub", "married", "profession", "company", "responsible_role", "engineer_specialty", "phone", "email", "facebook_url", "instagram_url", "viber", "city", "district", "address", "note"]
    w_headers, w_getters = _mn_headers_and_getters(Worker, worker_fields)
    _write_sheet(wb, "Ажиллагсад", w_headers, [[g(w) for g in w_getters] for w in workers])
    fam_fields = ["worker", "relation_type", "last_name", "first_name", "register_no", "birth_date", "phone", "email", "facebook_url", "instagram_url", "viber", "note"]
    f_headers, f_getters = _mn_headers_and_getters(FamilyMember, fam_fields)
    _write_sheet(wb, "Хамаарал", f_headers, [[g(fm) for g in f_getters] for fm in families])
    company_fields = ["name", "register_no", "activity_type", "activity_direction", "activity_sub_direction", "city", "district", "address", "phone", "email", "website", "note"]
    comp_headers, comp_getters = _mn_headers_and_getters(Company, company_fields)
    _write_sheet(wb, "Компани", comp_headers, [[g(c) for g in comp_getters] for c in companies])
    brigade_headers = ["Бригадын нэр", "Үйл ажиллагааны чиглэл", "Дэд сонголт(ууд)", "Ахлагч", "Хамтарч ажилладаг компаниуд", "Тайлбар"]
    b_rows = []
    for b in brigades:
        try: companies_txt = ", ".join([str(c) for c in b.companies.all()])
        except Exception: companies_txt = ""
        b_rows.append([_fmt_cell_value(getattr(b, "name", "")), _fmt_cell_value(b.get_activity_directions_display() if hasattr(b, "get_activity_directions_display") else getattr(b, "activity_directions_csv", "")), _fmt_cell_value(getattr(b, "activity_sub_directions_csv", "")), _fmt_cell_value(getattr(b, "leader_worker", "")) if getattr(b, "leader_worker_id", None) else "", companies_txt, _fmt_cell_value(getattr(b, "note", ""))])
    _write_sheet(wb, "Бригад", brigade_headers, b_rows)
    _write_sheet(wb, "Бригадын гишүүд", ["Бригад", "Бригадын гишүүн (Ажилтан)"], [[_fmt_cell_value(x.brigade), _fmt_cell_value(x.worker)] for x in brigade_members])
    return _export_workbook_response(wb, f"workers_export_mn_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

export_selected_workers_mn_xlsx.short_description = "Export (MN XLSX) — Сонгогдсон ажиллагсад + бусад"


class RoleScopedAdmin(ImportExportModelAdmin, GlobalContainsSearchAdmin):
    class Media:
        css = {"all": ("registry/admin/import_export_fix.css",)}

    import_template_enabled = True
    import_template_builder = None
    import_template_filename = None
    import_template_title = None

    def _model_key(self): return self.model._meta.model_name.lower()
    def _viewer_allowed(self): return bool(VIEWER_ALLOWED_MODELS.get(self._model_key(), False))

    def has_module_permission(self, request):
        u = request.user
        if _is_admin_user(u): return True
        if _is_operator(u): return True
        if _is_viewer(u): return self._viewer_allowed()
        return False

    def has_view_permission(self, request, obj=None):
        u = request.user
        if _is_admin_user(u): return True
        if _is_operator(u): return True
        if _is_viewer(u): return self._viewer_allowed()
        return False

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_export_permission(self, request): return _is_admin_user(request.user)
    def has_import_permission(self, request): return _is_admin_user(request.user)

    def _import_template_url_name(self):
        return f"{self.model._meta.app_label}_{self.model._meta.model_name}_import_template"

    def get_import_template_filename(self):
        return self.import_template_filename or f"{self.model._meta.model_name}_import_template.xlsx"

    def get_import_template_title(self):
        return self.import_template_title or f"{self.model._meta.verbose_name_raw.title()} Import Template"

    def get_import_template_content(self, request):
        if callable(self.import_template_builder): return self.import_template_builder()
        rc = getattr(self, "resource_class", None)
        if rc:
            try:
                res = rc()
                headers = res.get_export_headers()
                return build_xlsx_from_headers(headers)
            except Exception: pass
        return build_xlsx_from_headers(default_model_headers(self.model))

    def get_urls(self):
        urls = super().get_urls()
        if not getattr(self, "import_template_enabled", True): return urls
        my = [path("import-template/", self.admin_site.admin_view(self.import_template_view), name=self._import_template_url_name())]
        return my + urls

    def import_template_view(self, request):
        if not _is_admin_user(request.user): return self.admin_site.login(request)
        content = self.get_import_template_content(request)
        filename = self.get_import_template_filename()
        resp = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


def build_location_choices():
    out = [("", "---------")]
    for code, label in CITY_CHOICES:
        if code == "UB": continue
        out.append((code, label))
    for d_code, d_label in UB_DISTRICT_CHOICES:
        if d_code: out.append((f"UB|{d_code}", f"Улаанбаатар - {d_label}"))
    return out

LOCATION_CHOICES = build_location_choices()

class LocationMixinForm(forms.ModelForm):
    location = forms.ChoiceField(label="Оршин суугаа газрын хаяг", required=False, choices=LOCATION_CHOICES)

    def _init_location(self):
        city = (getattr(self.instance, "city", "") or "").strip()
        district = (getattr(self.instance, "district", "") or "").strip()
        if city == "UB" and district: self.fields["location"].initial = f"UB|{district}"
        elif city: self.fields["location"].initial = city
        else: self.fields["location"].initial = ""
        if "city" in self.fields:
            self.fields["city"].required = False
            self.fields["city"].widget = forms.HiddenInput()
        if "district" in self.fields:
            self.fields["district"].required = False
            self.fields["district"].widget = forms.HiddenInput()

    def _apply_location_to_instance(self):
        val = (self.cleaned_data.get("location") or "").strip()
        if not val:
            self.instance.city = ""
            self.instance.district = ""
            return
        if val.startswith("UB|"):
            self.instance.city = "UB"
            self.instance.district = val.split("|", 1)[1]
        else:
            self.instance.city = val
            self.instance.district = ""

def _normalize_url(url):
    u = (url or "").strip()
    if not u: return ""
    if u.startswith("http://") or u.startswith("https://"): return u
    return "https://" + u

COMPANY_ACTIVITY_SUB_MAP = {
    "CONSTRUCTION": ["Барилга угсралтын үйл ажиллагаа", "Барилгын засвар, шинэчлэл", "Барилгын дотор заслын ажил"],
    "DESIGN": ["Барилгын зураг зохиогч", "Төлөвчин"],
    "COMM_SYSTEM": ["Холбоо, дохиолол, автоматжуулалтын системийн угсралт", "Галын дохиолол, унтраах системийн угсралт", "Хяналтын камер, хамгаалалтын системийн угсралт"],
    "ENGINEERING_NETWORK": ["Гадна төв магистрал шугам", "Гадна дулаан хангамжийн шугам сүлжээ угсралт", "Гадна ус хангамж, ариутгах татуугын шугам угсралт", "Гадна цахилгаан хангамжийн шугам сүлжээ угсралт"],
    "EXTERNAL_ROAD": ["Гадна зам, талбайн тохижилт", "Гадна тохижилт, ногоон байгууламж", "Хашаа, хашлага барих ажил"],
}

class CompanyAdminForm(LocationMixinForm):
    activity_direction = forms.ChoiceField(label="Үйл ажиллагааны чиглэл", required=False, choices=[("", "---------")] + list(COMPANY_ACTIVITY_DIRECTION_CHOICES), widget=forms.Select())
    activity_sub_direction = forms.ChoiceField(label="Дэд сонголт", required=False, choices=[("", "---------")], widget=forms.Select())

    class Meta:
        model = Company
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._init_location()
        selected = (self.data.get("activity_direction") or getattr(self.instance, "activity_direction", "") or "").strip()
        subs = COMPANY_ACTIVITY_SUB_MAP.get(selected, [])
        self.fields["activity_sub_direction"].choices = [("", "---------")] + [(x, x) for x in subs]

    def clean(self):
        cleaned = super().clean()
        main = (cleaned.get("activity_direction") or "").strip()
        sub = (cleaned.get("activity_sub_direction") or "").strip()
        if sub and sub not in COMPANY_ACTIVITY_SUB_MAP.get(main, []): cleaned["activity_sub_direction"] = ""
        if main == "DESIGN": cleaned["activity_type"] = Company.ActivityType.DESIGN
        elif main in ("CONSTRUCTION", "ELECTRICAL_INTERNAL", "PLUMBING_INTERNAL", "HVAC", "COMM_SYSTEM", "ENGINEERING_NETWORK", "EXTERNAL_ROAD"): cleaned["activity_type"] = Company.ActivityType.CONSTRUCTION
        elif main in ("MATERIAL_PRODUCTION", "MATERIAL_TRADE"): cleaned["activity_type"] = Company.ActivityType.SUPPLY
        elif main == "SUPERVISION": cleaned["activity_type"] = Company.ActivityType.CONSULTING
        else: cleaned["activity_type"] = Company.ActivityType.OTHER
        return cleaned

    def save(self, commit=True):
        self._apply_location_to_instance()
        return super().save(commit=commit)


@admin.register(Company)
class CompanyAdmin(RoleScopedAdmin):
    form = CompanyAdminForm
    actions = [send_email_action, export_selected_companies_mn_xlsx]
    resource_class = CompanyResource
    import_template_builder = staticmethod(build_company_template_xlsx)
    import_template_filename = "company_import_template_mn.xlsx"
    import_template_title = "Company Import Template"

    def website_link(self, obj):
        u = _normalize_url(getattr(obj, "website", ""))
        if not u: return "-"
        return format_html('<a href="{}" target="_blank" rel="noopener noreferrer">{}</a>', u, u)
    website_link.short_description = "Вэб"
    website_link.admin_order_field = "website"

    def web_page_link(self, obj):
        if not obj.slug:
            return format_html('<span style="color:#aaa;font-size:12px;">slug алга</span>')
        url = f"/company/{obj.slug}/"
        return format_html(
            '<a href="{}" target="_blank" style="display:inline-flex;align-items:center;gap:4px;'
            'background:#f0a500;color:#0d1117;padding:3px 10px;border-radius:5px;'
            'font-size:12px;font-weight:700;text-decoration:none;">🌐 Web</a>',
            url
        )
    web_page_link.short_description = "Web хуудас"

    list_display = ("name", "activity_direction", "city", "phone", "email", "web_page_link", "website_link")
    list_filter = ("activity_direction", "city", "district")
    ordering = ("name",)
    fields = ("name", "slug", "register_no", "activity_direction", "activity_sub_direction", "location", "address", "phone", "email", "website", "facebook_url", "established_year", "employee_count", "description", "logo", "logo_url", "cover", "note")
    change_form_template = "admin/registry/company/change_form.html"

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        import json
        extra_context = extra_context or {}
        extra_context["COMPANY_ACTIVITY_SUB_MAP"] = json.dumps(COMPANY_ACTIVITY_SUB_MAP, ensure_ascii=False)
        # Web хуудасны URL-г template-д дамжуулна
        if request.resolver_match and request.resolver_match.kwargs.get("object_id"):
            obj_id = request.resolver_match.kwargs["object_id"]
            try:
                from .models import Company as _C
                obj = _C.objects.get(pk=obj_id)
                if obj.slug:
                    extra_context["company_web_url"] = f"/company/{obj.slug}/"
            except Exception:
                pass
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        u = request.user
        if _is_admin_user(u): return qs
        c = _user_company(u)
        if not c: return qs.none()
        return qs.filter(pk=c.pk)

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


def build_birth_sub_choices(city_code):
    out = [("", "---------")]
    if not city_code: return out
    if city_code == "UB":
        district_label = dict(UB_DISTRICT_CHOICES)
        for d_code in BIRTH_SOUMS.get("UB", []):
            out.append((d_code, district_label.get(d_code, d_code)))
        return out
    for name in BIRTH_SOUMS.get(city_code, []):
        out.append((name, name))
    return out


class WorkerAdminForm(LocationMixinForm):
    birth_place_city = forms.ChoiceField(label="Төрсөн газар - Аймаг/Хот", required=False, choices=[("", "---------")] + list(CITY_CHOICES), widget=forms.Select())
    birth_place_sub = forms.ChoiceField(label="Төрсөн газар - Сум/Дүүрэг", required=False, choices=[("", "---------")], widget=forms.Select())

    class Meta:
        model = Worker
        fields = "__all__"
        widgets = {"birth_date": forms.DateInput(attrs={"type": "date"})}

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._init_location()
        selected_city = (self.data.get("birth_place_city") or getattr(self.instance, "birth_place_city", "") or "").strip()
        self.fields["birth_place_sub"].choices = build_birth_sub_choices(selected_city)

    def clean(self):
        cleaned = super().clean()
        role = (cleaned.get("responsible_role") or "").strip()
        specialty = (cleaned.get("engineer_specialty") or "").strip()
        if role == "ENGINEER" and not specialty:
            self.add_error("engineer_specialty", "Инженер сонгосон тул 'Инженерийн төрөл' заавал сонгоно.")
        if role != "ENGINEER":
            cleaned["engineer_specialty"] = ""
            self.instance.engineer_specialty = ""
        return cleaned

    def save(self, commit=True):
        self._apply_location_to_instance()
        return super().save(commit=commit)


class BirthPlaceSubCountFilter(admin.SimpleListFilter):
    title = "Төрсөн сум/дүүрэг (тоо)"
    parameter_name = "birth_sub_count"

    def lookups(self, request, modeladmin):
        qs = Worker.objects.all()
        city = request.GET.get("birth_place_city") or ""
        if city: qs = qs.filter(birth_place_city=city)
        counts = qs.exclude(birth_place_sub="").values("birth_place_sub").annotate(c=Count("id")).order_by("-c", "birth_place_sub")
        return [(row["birth_place_sub"], f'{row["birth_place_sub"]} ({row["c"]})') for row in counts]

    def queryset(self, request, queryset):
        val = request.GET.get(self.parameter_name)
        if val: return queryset.filter(birth_place_sub=val)
        return queryset


class FamilyMemberAdminForm(forms.ModelForm):
    class Meta:
        model = FamilyMember
        fields = "__all__"
        widgets = {"birth_date": forms.DateInput(attrs={"type": "date"})}


class FamilyMemberInline(admin.TabularInline):
    model = FamilyMember
    form = FamilyMemberAdminForm
    extra = 0
    show_change_link = True
    fields = ("relation_type", "last_name", "first_name", "register_no", "birth_date", "phone", "email", "facebook_url", "viber", "note")


@admin.register(Worker)
class WorkerAdmin(RoleScopedAdmin):
    form = WorkerAdminForm
    actions = [send_email_action, export_selected_workers_mn_xlsx]
    inlines = [FamilyMemberInline]
    resource_class = WorkerResource
    import_template_builder = staticmethod(build_worker_template_xlsx)
    import_template_filename = "worker_import_template_mn.xlsx"
    import_template_title = "Worker Import Template"

    list_display = ("last_name", "parent_name", "first_name", "gender", "register_no", "birth_date", "birth_place_city", "birth_place_sub", "married", "profession", "company", "responsible_role", "engineer_specialty", "phone", "email")
    ordering = ("first_name", "last_name")
    autocomplete_fields = ("company",)
    list_filter = ("gender", "married", "birth_place_city", BirthPlaceSubCountFilter, "profession", "company", "city", "district")
    fields = ("last_name", "parent_name", "first_name", "gender", "register_no", "birth_date", "birth_place_city", "birth_place_sub", "married", "profession", "company", "responsible_role", "engineer_specialty", "phone", "email", "facebook_url", "viber", "instagram_url", "location", "address", "note")
    change_form_template = "admin/registry/worker/change_form.html"

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        import json
        extra_context = extra_context or {}
        extra_context["BIRTH_SOUMS"] = json.dumps(BIRTH_SOUMS, ensure_ascii=False)
        return super().changeform_view(request, object_id, form_url, extra_context=extra_context)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        u = request.user
        if _is_admin_user(u): return qs
        c = _user_company(u)
        if not c: return qs.none()
        return qs.filter(company=c)

    def has_add_permission(self, request):
        u = request.user
        if _is_admin_user(u): return True
        return _is_operator(u)

    def has_change_permission(self, request, obj=None):
        u = request.user
        if _is_admin_user(u): return True
        if not _is_operator(u): return False
        if obj is None: return True
        c = _user_company(u)
        return bool(c and getattr(obj, "company_id", None) == c.id)

    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj=obj, change=change, **kwargs)
        u = request.user
        if _is_admin_user(u): return form
        if _is_operator(u):
            c = _user_company(u)
            if c and "company" in form.base_fields:
                form.base_fields["company"].queryset = Company.objects.filter(pk=c.pk)
                form.base_fields["company"].initial = c.pk
        return form

    def save_model(self, request, obj, form, change):
        u = request.user
        if _is_admin_user(u): return super().save_model(request, obj, form, change)
        if _is_operator(u):
            c = _user_company(u)
            if c: obj.company = c
        return super().save_model(request, obj, form, change)


@admin.register(GovernmentOrganization)
class GovernmentOrganizationAdmin(RoleScopedAdmin):
    import_template_builder = staticmethod(build_government_org_template_xlsx)
    import_template_filename = "government_org_import_template_mn.xlsx"
    import_template_title = "Салбарын төрийн байгууллага MN загвар"

    def website_link(self, obj):
        u = _normalize_url(getattr(obj, "website", ""))
        if not u: return "-"
        return format_html('<a href="{}" target="_blank" rel="noopener noreferrer">{}</a>', u, u)
    website_link.short_description = "Вэб"
    website_link.admin_order_field = "website"

    list_display = ("name", "register_no", "phone", "email", "website_link")
    list_filter = ("name", "register_no")
    ordering = ("name",)

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(NonGovernmentOrganization)
class NonGovernmentOrganizationAdmin(RoleScopedAdmin):
    import_template_builder = staticmethod(build_non_government_org_template_xlsx)
    import_template_filename = "ngo_import_template_mn.xlsx"
    import_template_title = "Салбарын төрийн бус байгууллага MN загвар"

    def website_link(self, obj):
        u = _normalize_url(getattr(obj, "website", ""))
        if not u: return "-"
        return format_html('<a href="{}" target="_blank" rel="noopener noreferrer">{}</a>', u, u)
    website_link.short_description = "Вэб"
    website_link.admin_order_field = "website"

    list_display = ("name", "register_no", "phone", "email", "website_link")
    list_filter = ("name", "register_no")
    ordering = ("name",)

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(FamilyMember)
class FamilyMemberAdmin(RoleScopedAdmin):
    form = FamilyMemberAdminForm
    autocomplete_fields = ("worker",)
    import_template_builder = staticmethod(build_family_member_template_xlsx)
    import_template_filename = "family_member_import_template_mn.xlsx"
    import_template_title = "Ажилтны хамаарал MN загвар"

    list_display = ("last_name", "first_name", "register_no", "relation_type", "worker", "birth_date", "phone", "email")
    ordering = ("worker", "first_name", "last_name")
    list_filter = ("relation_type", "worker", "birth_date")
    fields = ("worker", "relation_type", "last_name", "first_name", "register_no", "birth_date", "phone", "email", "facebook_url", "instagram_url", "viber", "note", "full_name")
    readonly_fields = ("full_name",)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        u = request.user
        if _is_admin_user(u): return qs
        c = _user_company(u)
        if not c: return qs.none()
        return qs.filter(worker__company=c)

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


def _worker_payload(w):
    if not w:
        return {"ok": False, "id": None, "register_no": "", "full_name": "", "gender": "", "birth_date": "", "birth_place_city": "", "birth_place_sub": "", "married": "", "profession": "", "company": "", "responsible_role": "", "engineer_specialty": "", "phone": "", "email": "", "facebook_url": "", "instagram_url": "", "viber": "", "city": "", "district": "", "address": ""}
    full = " ".join(f"{w.last_name} {w.parent_name} {w.first_name}".split()).strip()
    return {
        "ok": True, "id": w.id, "register_no": w.register_no or "", "full_name": full,
        "gender": w.get_gender_display() if getattr(w, "gender", "") else "",
        "birth_date": (w.birth_date.isoformat() if getattr(w, "birth_date", None) else ""),
        "birth_place_city": w.get_birth_place_city_display() if getattr(w, "birth_place_city", "") else "",
        "birth_place_sub": (w.birth_place_sub or ""),
        "married": ("Тийм" if getattr(w, "married", False) else "Үгүй"),
        "profession": w.get_profession_display() if getattr(w, "profession", "") else "",
        "company": (str(w.company) if getattr(w, "company", None) else ""),
        "responsible_role": w.get_responsible_role_display() if getattr(w, "responsible_role", "") else "",
        "engineer_specialty": w.get_engineer_specialty_display() if getattr(w, "engineer_specialty", "") else "",
        "phone": w.phone or "", "email": w.email or "",
        "facebook_url": getattr(w, "facebook_url", "") or "",
        "instagram_url": getattr(w, "instagram_url", "") or "",
        "viber": getattr(w, "viber", "") or "",
        "city": w.get_city_display() if getattr(w, "city", "") else "",
        "district": w.get_district_display() if getattr(w, "district", "") else "",
        "address": getattr(w, "address", "") or "",
    }

def _brigade_allowed_subs(main_codes):
    allowed = []
    for c in (main_codes or []):
        allowed += COMPANY_ACTIVITY_SUB_MAP.get(c, [])
    seen = set()
    out = []
    for x in allowed:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


class BrigadeActivityFilter(admin.SimpleListFilter):
    title = "Үйл ажиллагааны чиглэл"
    parameter_name = "brigade_activity"

    def lookups(self, request, modeladmin): return list(COMPANY_ACTIVITY_DIRECTION_CHOICES)

    def queryset(self, request, queryset):
        code = request.GET.get(self.parameter_name) or ""
        if not code: return queryset
        return queryset.filter(activity_directions_csv__contains=f",{code},")


class BrigadeAdminForm(forms.ModelForm):
    activity_directions = forms.MultipleChoiceField(label="Үйл ажиллагааны чиглэл", required=False, choices=COMPANY_ACTIVITY_DIRECTION_CHOICES, widget=forms.CheckboxSelectMultiple())
    activity_sub_directions = forms.MultipleChoiceField(label="Үйл ажиллагааны дэд сонголт", required=False, choices=[], widget=forms.SelectMultiple(attrs={"size": "8"}))
    leader_register_no_display = forms.CharField(label="Регистр", required=False, disabled=True)
    leader_full_name_display = forms.CharField(label="Овог нэр", required=False, disabled=True)
    leader_gender_display = forms.CharField(label="Хүйс", required=False, disabled=True)
    leader_birth_date_display = forms.CharField(label="Төрсөн огноо", required=False, disabled=True)
    leader_birth_place_display = forms.CharField(label="Төрсөн газар", required=False, disabled=True)
    leader_married_display = forms.CharField(label="Гэрлэсэн эсэх", required=False, disabled=True)
    leader_profession_display = forms.CharField(label="Мэргэжил", required=False, disabled=True)
    leader_company_display = forms.CharField(label="Харьяалах компани", required=False, disabled=True)
    leader_role_display = forms.CharField(label="Хариуцсан ажил", required=False, disabled=True)
    leader_specialty_display = forms.CharField(label="Инженерийн төрөл", required=False, disabled=True)
    leader_phone_display = forms.CharField(label="Утас", required=False, disabled=True)
    leader_email_display = forms.CharField(label="Имэйл", required=False, disabled=True)
    leader_social_display = forms.CharField(label="Сошиал", required=False, disabled=True)
    leader_address_display = forms.CharField(label="Оршин суугаа", required=False, disabled=True)

    class Meta:
        model = Brigade
        exclude = ("activity_directions_csv", "activity_sub_directions_csv")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.data: main_codes = self.data.getlist("activity_directions")
        else: main_codes = self.instance.get_activity_direction_codes()
        self.fields["activity_directions"].initial = main_codes
        allowed_subs = _brigade_allowed_subs(main_codes)
        self.fields["activity_sub_directions"].choices = [(x, x) for x in allowed_subs]
        if self.data: sub_initial = self.data.getlist("activity_sub_directions")
        else: sub_initial = self.instance.get_activity_sub_codes()
        self.fields["activity_sub_directions"].initial = [x for x in sub_initial if x in allowed_subs]

    def clean(self):
        cleaned = super().clean()
        mains = cleaned.get("activity_directions") or []
        subs = cleaned.get("activity_sub_directions") or []
        allowed = set(_brigade_allowed_subs(mains))
        cleaned["activity_sub_directions"] = [s for s in subs if s in allowed]
        return cleaned

    def save(self, commit=True):
        mains = self.cleaned_data.get("activity_directions") or []
        subs = self.cleaned_data.get("activity_sub_directions") or []
        self.instance.set_activity_direction_codes(mains)
        self.instance.set_activity_sub_codes(subs)
        return super().save(commit=commit)


class BrigadeMemberInlineForm(forms.ModelForm):
    worker_register_no_display = forms.CharField(label="Регистр", required=False, disabled=True)
    worker_full_name_display = forms.CharField(label="Овог нэр", required=False, disabled=True)
    worker_phone_display = forms.CharField(label="Утас", required=False, disabled=True)
    worker_email_display = forms.CharField(label="Имэйл", required=False, disabled=True)
    worker_company_display = forms.CharField(label="Компани", required=False, disabled=True)

    class Meta:
        model = BrigadeMember
        fields = "__all__"


class BrigadeMemberInline(admin.TabularInline):
    model = BrigadeMember
    form = BrigadeMemberInlineForm
    extra = 0
    show_change_link = False
    autocomplete_fields = ("worker",)
    fields = ("worker", "worker_register_no_display", "worker_full_name_display", "worker_phone_display", "worker_email_display", "worker_company_display")


@admin.register(Brigade)
class BrigadeAdmin(RoleScopedAdmin):
    form = BrigadeAdminForm
    inlines = [BrigadeMemberInline]
    autocomplete_fields = ("companies", "leader_worker")
    ordering = ("name",)
    import_template_builder = staticmethod(build_brigade_template_xlsx)
    import_template_filename = "brigade_import_template_mn.xlsx"
    import_template_title = "Барилгын бригад MN загвар"

    def activities_display(self, obj): return obj.get_activity_directions_display()
    activities_display.short_description = "Үйл ажиллагааны чиглэл"

    list_display = ("name", "activities_display", "leader_worker", "member_count")
    list_filter = (BrigadeActivityFilter, "companies")

    fieldsets = (
        ("Бригад", {"fields": ("name", "activity_directions", "activity_sub_directions", "companies", "note")}),
        ("Бригадын ахлагч", {"fields": ("leader_worker",)}),
        ("Ахлагчийн мэдээлэл (сонгоход автоматаар)", {"fields": ("leader_register_no_display", "leader_full_name_display", "leader_gender_display", "leader_birth_date_display", "leader_birth_place_display", "leader_married_display", "leader_profession_display", "leader_company_display", "leader_role_display", "leader_specialty_display", "leader_phone_display", "leader_email_display", "leader_social_display", "leader_address_display")}),
    )

    class Media:
        js = ("registry/admin/brigade_autofill.js", "registry/admin/brigade_activity_sub.js")

    def get_urls(self):
        urls = super().get_urls()
        my = [
            path("worker-info/<int:pk>/", self.admin_site.admin_view(self.worker_info_view), name="registry_brigade_worker_info"),
            path("sub-map/", self.admin_site.admin_view(self.sub_map_view), name="registry_brigade_sub_map"),
        ]
        return my + urls

    def worker_info_view(self, request, pk):
        w = Worker.objects.filter(pk=pk).first()
        return JsonResponse(_worker_payload(w))

    def sub_map_view(self, request): return JsonResponse(COMPANY_ACTIVITY_SUB_MAP)

    def member_count(self, obj): return obj.members.count()
    member_count.short_description = "Гишүүд"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        u = request.user
        if _is_admin_user(u): return qs
        c = _user_company(u)
        if not c: return qs.none()
        return qs.filter(companies=c).distinct()

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(BrigadeMember)
class BrigadeMemberAdmin(RoleScopedAdmin):
    import_template_builder = staticmethod(build_brigade_member_template_xlsx)
    import_template_filename = "brigade_member_import_template_mn.xlsx"
    import_template_title = "Бригадын гишүүн MN загвар"

    autocomplete_fields = ("brigade", "worker")
    list_display = ("brigade", "worker")
    list_filter = ("brigade",)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        u = request.user
        if _is_admin_user(u): return qs
        c = _user_company(u)
        if not c: return qs.none()
        return qs.filter(worker__company=c)

    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(AdminGroup)
class AdminGroupAdmin(RoleScopedAdmin):
    list_display = ("name",)
    ordering = ("name",)

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(UserCompanyProfile)
class UserCompanyProfileAdmin(RoleScopedAdmin):
    import_template_enabled = False
    list_display = ("user", "company")
    autocomplete_fields = ("user", "company")
    search_fields = ("user__username", "user__email", "company__name")

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


User = get_user_model()
try: admin.site.unregister(User)
except admin.sites.NotRegistered: pass
try: admin.site.unregister(Group)
except admin.sites.NotRegistered: pass


@admin.register(User)
class UserAdmin(DjangoUserAdmin):
    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(Group)
class GroupAdmin(DjangoGroupAdmin):
    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(Banner)
class BannerAdmin(RoleScopedAdmin):
    import_template_enabled = False
    list_display = ("title", "is_active", "sort_order", "created_at", "image_url", "link_url")
    list_filter = ("is_active",)
    search_fields = ("title", "image_url", "link_url")
    ordering = ("-is_active", "sort_order", "-created_at")

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(PublicPost)
class PublicPostAdmin(RoleScopedAdmin):
    import_template_enabled = False
    list_display = ("title", "author", "is_published", "created_at")
    list_filter = ("is_published", "created_at")
    search_fields = ("title", "body", "author__username", "author__email")
    ordering = ("-created_at",)

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(HeroBanner)
class HeroBannerAdmin(admin.ModelAdmin):
    import_template_enabled = False
    list_display = ("title", "media_type", "is_active", "sort_order", "created_at")
    list_filter = ("is_active", "media_type")
    ordering = ("sort_order", "-created_at")

    fieldsets = (
        ("Үндсэн мэдээлэл", {"fields": ("title", "subtitle", "is_active", "sort_order")}),
        ("Медиа (зураг эсвэл видео)", {"fields": ("media_type", "image", "image_url", "video", "video_url"), "description": "Зүүн 1/3 хэсэгт харагдах медиа."}),
        ("Товчнууд", {"fields": ("btn1_text", "btn1_url", "btn2_text", "btn2_url")}),
    )

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(SliderAd)
class SliderAdAdmin(admin.ModelAdmin):
    import_template_enabled = False
    list_display = ("title", "description", "is_active", "sort_order", "created_at")
    list_filter = ("is_active",)
    ordering = ("sort_order", "-created_at")

    fieldsets = (
        ("Үндсэн мэдээлэл", {"fields": ("title", "description", "is_active", "sort_order")}),
        ("Зураг", {"fields": ("image", "image_url")}),
        ("Холбоос", {"fields": ("link_url",)}),
    )

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


@admin.register(SubBanner)
class SubBannerAdmin(admin.ModelAdmin):
    import_template_enabled = False
    list_display = ("title", "is_active", "sort_order", "created_at")
    list_filter = ("is_active",)
    ordering = ("sort_order", "-created_at")

    fieldsets = (
        ("Үндсэн мэдээлэл", {"fields": ("title", "is_active", "sort_order")}),
        ("Зураг", {"fields": ("image", "image_url")}),
        ("Холбоос", {"fields": ("link_url",)}),
    )

    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_add_permission(self, request): return _is_admin_user(request.user)
    def has_change_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)


def _build_block(model, title, qs, request, limit=30):
    items = []
    for obj in qs[:limit]:
        try: url = reverse(f"admin:{obj._meta.app_label}_{obj._meta.model_name}_change", args=[obj.pk])
        except Exception: url = "#"
        items.append({"id": obj.pk, "label": str(obj), "change_url": url})
    return {"title": title, "count": qs.count(), "items": items}


def global_admin_search_view(request):
    if not _is_admin_user(request.user): return admin.site.login(request)
    q = (request.GET.get("q") or "").strip()
    results = []
    if q:
        models_to_search = [
            (Worker, "Ажиллагсад"), (FamilyMember, "Ажилтны хамаарал"),
            (Company, "Компани"), (Brigade, "Бригад"), (BrigadeMember, "Бригадын гишүүн"),
            (GovernmentOrganization, "Төрийн байгууллага"), (NonGovernmentOrganization, "ТББ"),
            (Banner, "Баннер"), (PublicPost, "Пост"),
        ]
        for model, title in models_to_search:
            qs = model.objects.filter(global_search_q(model, q))
            if qs.exists(): results.append(_build_block(model, title, qs, request))
    context = dict(admin.site.each_context(request), q=q, results=results, title="Нэгдсэн хайлт")
    return TemplateResponse(request, "admin/global_search.html", context)


_old_get_urls = admin.site.get_urls

def _new_get_urls():
    urls = _old_get_urls()
    custom = [path("global-search/", admin.site.admin_view(global_admin_search_view), name="global-search")]
    return custom + urls

admin.site.get_urls = _new_get_urls


# =====================================================
# ✅ SiteConfig Admin — Системийн тохиргоо
# =====================================================
@admin.register(SiteConfig)
class SiteConfigAdmin(admin.ModelAdmin):

    fieldsets = (
        ("📧 Email (SMTP) тохиргоо", {
            "fields": ("sender_name", "sender_email", "sender_phone", "email_host", "email_port", "email_use_tls", "email_host_user", "email_host_password"),
            "description": "Gmail: host=smtp.gmail.com, port=587, TLS=тийм. App Password: Google → Аюулгүй байдал → 2FA → App passwords",
        }),
        ("📱 SMS тохиргоо", {
            "fields": ("sms_gateway_url", "sms_gateway_token", "sms_sender_name"),
        }),
        ("✈️ Telegram тохиргоо", {
            "fields": ("telegram_bot_token",),
            "description": "@BotFather-с bot үүсгэж token авна.",
        }),
        ("📘 Facebook тохиргоо", {
            "fields": ("facebook_page_token",),
        }),
        ("💬 Viber тохиргоо", {
            "fields": ("viber_auth_token",),
        }),
        ("🌐 Сайтын мэдээлэл", {
            "fields": ("site_name", "site_phone", "site_email", "site_address", "site_facebook"),
        }),
    )

    def has_add_permission(self, request):
        return not SiteConfig.objects.exists() and _is_admin_user(request.user)

    def has_change_permission(self, request, obj=None):
        return _is_admin_user(request.user)

    def has_delete_permission(self, request, obj=None):
        return False

    def has_module_permission(self, request):
        return _is_admin_user(request.user)

    def has_view_permission(self, request, obj=None):
        return _is_admin_user(request.user)

    def changelist_view(self, request, extra_context=None):
        obj, _ = SiteConfig.objects.get_or_create(pk=1)
        from django.shortcuts import redirect
        from django.urls import reverse
        return redirect(reverse("admin:registry_siteconfig_change", args=[obj.pk]))


# =====================================================
# ✅ MessageLog Admin — Мессежийн бүртгэл
# =====================================================
@admin.register(MessageLog)
class MessageLogAdmin(admin.ModelAdmin):
    list_display = ("created_at", "channel", "recipient_name", "recipient_address", "subject_short", "status", "sent_by")
    list_filter = ("channel", "status", "created_at")
    search_fields = ("recipient_name", "recipient_address", "subject", "body")
    ordering = ("-created_at",)
    readonly_fields = ("sent_by", "channel", "target_type", "recipient_name", "recipient_address", "subject", "body", "status", "error_message", "created_at")

    def subject_short(self, obj):
        return (obj.subject or obj.body or "")[:50]
    subject_short.short_description = "Гарчиг/Агуулга"

    def changelist_view(self, request, extra_context=None):
        from apps.registry.admin_messaging_view import messaging_admin_view
        # Зөвхөн MessageLog-ийн changelist дээр ажиллана
        if self.model._meta.model_name == 'messagelog':
            return messaging_admin_view(request)
        return super().changelist_view(request, extra_context)

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None): return _is_admin_user(request.user)
    def has_module_permission(self, request): return _is_admin_user(request.user)
    def has_view_permission(self, request, obj=None): return _is_admin_user(request.user)
