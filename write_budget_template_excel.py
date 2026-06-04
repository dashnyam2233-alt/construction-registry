import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Материалын жагсаалт"

# Styles
header_fill = PatternFill("solid", fgColor="1e3a4a")
header_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="f59e0b")
section_font = Font(bold=True, color="1e3a4a", size=11)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center")

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 20

# Title
ws.merge_cells("A1:F1")
ws["A1"] = "БАРИЛГЫН МАТЕРИАЛ, АЖЛЫН ХЭМЖЭЭНИЙ ЖАГСААЛТ"
ws["A1"].font = Font(bold=True, size=14, color="1e3a4a")
ws["A1"].alignment = center
ws["A1"].fill = PatternFill("solid", fgColor="fef3c7")

ws.merge_cells("A2:F2")
ws["A2"] = "Барилгын нэр: _______________  Байршил: _______________  Огноо: _______________"
ws["A2"].font = Font(italic=True, size=10, color="64748b")
ws["A2"].alignment = Alignment(horizontal="left")
ws.row_dimensions[2].height = 20

# Header
headers = ["№", "Материал / Ажлын төрөл", "Нэгж", "Тоо хэмжээ", "Нэгж үнэ (₮)", "Нийт (₮)"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border
ws.row_dimensions[3].height = 22

# Sections
sections = [
    ("СУУРИЙН АЖИЛ", [
        ("Газар шорооны ажил (ухалт)", "м³", ""),
        ("Хэв хашмал угсралт", "м²", ""),
        ("Арматур угсралт", "тн", ""),
        ("Бетон цутгалт (суурь)", "м³", ""),
        ("Гидроизоляц", "м²", ""),
    ]),
    ("ХАНА, КАРКАС", [
        ("Тоосго / Блок өрөх", "м²", ""),
        ("Арматур угсралт (хана)", "тн", ""),
        ("Бетон цутгалт (хана)", "м³", ""),
        ("Дотор хуваалт", "м²", ""),
    ]),
    ("ДЭЭВЭР", [
        ("Дээврийн каркас угсралт", "м²", ""),
        ("Дээврийн материал", "м²", ""),
        ("Дулаалга", "м²", ""),
        ("Ус тусгаарлалт", "м²", ""),
    ]),
    ("ЦОН, ХААЛГА", [
        ("Гадна хаалга", "ш", ""),
        ("Дотор хаалга", "ш", ""),
        ("Цонх", "ш", ""),
    ]),
    ("ДОТОР ЗАСАЛ", [
        ("Шавардлага (хана)", "м²", ""),
        ("Плита наах (хана)", "м²", ""),
        ("Эмульс будаг", "м²", ""),
        ("Шал (ламинат/плита/паркет)", "м²", ""),
        ("Тааз (гипрок/будаг)", "м²", ""),
    ]),
    ("САНТЕХНИК", [
        ("Шугам хоолой (PPR)", "м", ""),
        ("Угаалтуур", "ш", ""),
        ("Суултуур", "ш", ""),
        ("Ванн / Душ", "ш", ""),
        ("Радиатор", "ш", ""),
    ]),
    ("ЦАХИЛГААН", [
        ("Цахилгааны кабель", "м", ""),
        ("Унтраалга", "ш", ""),
        ("Розетка", "ш", ""),
        ("Гэрэлтүүлэг", "ш", ""),
    ]),
    ("АЖИЛЧИД", [
        ("Барилгачин (ерөнхий)", "хүн/өдөр", ""),
        ("Мэргэжилтэн (цахилгаан, сантехник)", "хүн/өдөр", ""),
        ("Инженер хяналт", "сар", ""),
    ]),
    ("ТЭЭВЭР, МАШИН", [
        ("Материал тээвэр", "удаа", ""),
        ("Кран үйлчилгээ", "цаг", ""),
        ("Экскаватор", "цаг", ""),
        ("Хог зайлуулах", "удаа", ""),
    ]),
    ("БУСАД ЗАРДАЛ", [
        ("Зураг төсөл", "удаа", ""),
        ("Барилгын зөвшөөрөл", "удаа", ""),
        ("Барилгын даатгал", "жил", ""),
        ("НӨАТ (10%)", "хувь", "10"),
    ]),
]

row = 4
num = 1
for section_name, items in sections:
    # Section header
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = section_name
    ws[f"A{row}"].fill = section_fill
    ws[f"A{row}"].font = section_font
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws[f"A{row}"].border = border
    ws.row_dimensions[row].height = 20
    row += 1

    for name, unit, default in items:
        ws.cell(row=row, column=1, value=num).alignment = center
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=name).border = border
        ws.cell(row=row, column=3, value=unit).alignment = center
        ws.cell(row=row, column=3).border = border
        qty_cell = ws.cell(row=row, column=4, value=float(default) if default else None)
        qty_cell.alignment = center
        qty_cell.border = border
        qty_cell.number_format = "#,##0.00"
        price_cell = ws.cell(row=row, column=5, value=None)
        price_cell.alignment = Alignment(horizontal="right")
        price_cell.border = border
        price_cell.number_format = "#,##0"
        price_cell.fill = PatternFill("solid", fgColor="fffbeb")
        # Total formula
        total_cell = ws.cell(row=row, column=6)
        total_cell.value = f"=IF(AND(D{row}<>\"\",E{row}<>\"\"),D{row}*E{row},\"\")"
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.border = border
        total_cell.number_format = "#,##0"
        ws.row_dimensions[row].height = 18
        num += 1
        row += 1

# Grand total
ws.merge_cells(f"A{row}:E{row}")
ws[f"A{row}"] = "НИЙТ ТӨСӨВ"
ws[f"A{row}"].font = Font(bold=True, size=13, color="1e3a4a")
ws[f"A{row}"].fill = PatternFill("solid", fgColor="f59e0b")
ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
ws[f"A{row}"].border = border
ws[f"F{row}"] = f"=SUM(F4:F{row-1})"
ws[f"F{row}"].number_format = "#,##0"
ws[f"F{row}"].font = Font(bold=True, size=13, color="1e3a4a")
ws[f"F{row}"].fill = PatternFill("solid", fgColor="f59e0b")
ws[f"F{row}"].alignment = Alignment(horizontal="right")
ws[f"F{row}"].border = border
ws.row_dimensions[row].height = 26

# Note
row += 2
ws.merge_cells(f"A{row}:F{row}")
ws[f"A{row}"] = "⚠️ Тайлбар: Шар өнгийн нүдэнд нэгж үнийг оруулна уу. Тоо хэмжээ, нэгж үнэ оруулсны дараа нийт автоматаар тооцогдоно."
ws[f"A{row}"].font = Font(italic=True, size=10, color="854d0e")
ws[f"A{row}"].fill = PatternFill("solid", fgColor="fef9ec")

import os
os.makedirs("media/budget_templates", exist_ok=True)
wb.save("media/budget_templates/budget_template.xlsx")
print("OK — budget_template.xlsx үүслээ")