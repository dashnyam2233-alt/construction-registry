# Excel view нэмэх
content = open("apps/registry/views.py", "r", encoding="utf-8").read()

excel_view = '''

def budget_excel(request):
    import json
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    data_str = request.GET.get("data", "{}")
    try:
        data = json.loads(data_str)
    except:
        return HttpResponse("Алдаа", status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Төсөв"

    # Styles
    header_fill = PatternFill("solid", fgColor="1e3a4a")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill("solid", fgColor="f59e0b")
    section_font = Font(bold=True, color="1e3a4a", size=11)
    total_fill = PatternFill("solid", fgColor="fef3c7")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    row = 1

    # Title
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "БАРИЛГЫН ТӨСВИЙН ТООЦОО"
    ws[f"A{row}"].font = Font(bold=True, size=14, color="1e3a4a")
    ws[f"A{row}"].alignment = center
    row += 1

    # Building info
    info = data.get("building_info", {})
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"{info.get('type','')} | {info.get('area','')} | {info.get('location','')} | {info.get('quality','')}"
    ws[f"A{row}"].font = Font(italic=True, color="64748b")
    ws[f"A{row}"].alignment = center
    row += 2

    def write_section(title, items, total_key):
        nonlocal row
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        headers = ["Нэр", "Нэгж", "Тоо хэмжээ", "Нэгж үнэ (₮)", "Нийт (₮)"]
        aligns = ["left", "center", "right", "right", "right"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=i+1, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal=aligns[i], vertical="center")
            cell.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        # Data rows
        for item in items:
            ws.cell(row=row, column=1, value=item.get("name","")).border = border
            ws.cell(row=row, column=2, value=item.get("unit","")).alignment = center
            ws.cell(row=row, column=2).border = border
            ws.cell(row=row, column=3, value=item.get("qty",0)).alignment = right
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4, value=item.get("unit_price",0)).number_format = "#,##0"
            ws.cell(row=row, column=4).alignment = right
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=5, value=item.get("total",0)).number_format = "#,##0"
            ws.cell(row=row, column=5).alignment = right
            ws.cell(row=row, column=5).border = border
            row += 1

        # Total row
        summary = data.get("summary", {})
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "Нийт дүн"
        ws[f"A{row}"].fill = total_fill
        ws[f"A{row}"].font = total_font
        ws[f"A{row}"].border = border
        ws[f"E{row}"] = summary.get(total_key, 0)
        ws[f"E{row}"].number_format = "#,##0"
        ws[f"E{row}"].fill = total_fill
        ws[f"E{row}"].font = total_font
        ws[f"E{row}"].alignment = right
        ws[f"E{row}"].border = border
        row += 2

    write_section("🧱 МАТЕРИАЛЫН ЗАРДАЛ", data.get("materials", []), "materials_total")
    write_section("👷 АЖИЛЧДЫН ЗАРДАЛ", data.get("labor", []), "labor_total")
    write_section("📦 БУСАД ЗАРДАЛ", data.get("other", []), "other_total")

    # Grand total
    summary = data.get("summary", {})
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "НИЙТ ТӨСӨВ"
    ws[f"A{row}"].fill = PatternFill("solid", fgColor="f59e0b")
    ws[f"A{row}"].font = Font(bold=True, size=13, color="1e3a4a")
    ws[f"A{row}"].border = border
    ws[f"E{row}"] = summary.get("grand_total", 0)
    ws[f"E{row}"].number_format = "#,##0"
    ws[f"E{row}"].fill = PatternFill("solid", fgColor="f59e0b")
    ws[f"E{row}"].font = Font(bold=True, size=13, color="1e3a4a")
    ws[f"E{row}"].alignment = right
    ws[f"E{row}"].border = border
    ws.row_dimensions[row].height = 26
    row += 1

    # Per m2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "1 м² үнэ"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"E{row}"] = summary.get("price_per_m2", 0)
    ws[f"E{row}"].number_format = "#,##0"
    ws[f"E{row}"].alignment = right
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Барилгын хугацаа"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"E{row}"] = f"{summary.get('duration_months', 0)} сар"
    ws[f"E{row}"].alignment = right
    row += 2

    # Notes
    if data.get("notes"):
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = f"Анхаарах зүйлс: {data['notes']}"
        ws[f"A{row}"].font = Font(italic=True, color="166534")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="f0fdf4")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=budget.xlsx"
    wb.save(response)
    return response
'''

if "def budget_excel" not in content:
    with open("apps/registry/views.py", "a", encoding="utf-8") as f:
        f.write(excel_view)
    print("OK — excel view нэмэгдлээ")

# URL нэмэх
urls = open("apps/registry/urls.py", "r", encoding="utf-8").read()
if "budget_excel" not in urls:
    urls = urls.replace(
        "from .views import (\n    budget_calculator,",
        "from .views import (\n    budget_calculator,\n    budget_excel,"
    )
    urls = urls.replace(
        'path("budget/", budget_calculator, name="budget_calculator"),',
        'path("budget/", budget_calculator, name="budget_calculator"),\n    path("budget/excel/", budget_excel, name="budget_excel"),'
    )
    open("apps/registry/urls.py", "w", encoding="utf-8").write(urls)
    print("OK — URL нэмэгдлээ")

print("Дууслаа")